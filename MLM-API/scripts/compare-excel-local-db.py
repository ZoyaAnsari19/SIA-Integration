#!/usr/bin/env python3
"""
Compare Excel file (products-export-3.xlsx) with Local Database
Checks: Users, Packages, and their amounts
"""

import openpyxl
import subprocess
import json
from datetime import datetime

# Configuration
EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'  # Current local database
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        return None
    
    return result.stdout.strip()

def get_excel_users():
    """Read all users from Excel"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return {}
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    excel_users = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        user_id = ws.cell(row_idx, headers.get('ID', 0)).value if headers.get('ID') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
        
        if not email and not member_id:
            continue
        
        # Get packages
        packages = []
        for pack_num in range(1, 9):
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_name_col = f'Pack. {pack_num} name'
            
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            if pack_amt is not None:
                try:
                    amount = float(pack_amt) if isinstance(pack_amt, (int, float)) else None
                    if amount and amount > 0:
                        pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
                        packages.append({
                            'amount': amount,
                            'name': str(pack_name) if pack_name else 'N/A'
                        })
                except:
                    pass
        
        # Create key for matching (email or member_id)
        key = email.lower() if email else (member_id.upper() if member_id else None)
        if not key:
            continue
        
        excel_users[key] = {
            'email': email,
            'member_id': member_id,
            'name': name,
            'user_id': user_id,
            'packages': packages
        }
    
    print(f"✅ Found {len(excel_users)} users in Excel")
    return excel_users

def get_db_users_with_packages():
    """Get all users with their packages from database"""
    print("🗄️  Reading database...")
    
    # Get all users
    query = """
    SELECT 
        u.id,
        u.display_id,
        u.email,
        u.name,
        COUNT(DISTINCT p.id) as package_count
    FROM users u
    LEFT JOIN purchases p ON p.user_id = u.id
    GROUP BY u.id, u.display_id, u.email, u.name
    ORDER BY u.id;
    """
    
    result = run_sql_query(query)
    if not result:
        print("❌ Failed to get users from database")
        return {}
    
    db_users = {}
    
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 5:
            user_id = parts[0].strip()
            display_id = parts[1].strip() if parts[1] else None
            email = parts[2].strip() if parts[2] else None
            name = parts[3].strip() if parts[3] else None
            package_count = int(parts[4].strip()) if parts[4] else 0
            
            # Get packages for this user
            pkg_query = f"""
            SELECT 
                pkg.id,
                pkg.name,
                pkg.price,
                pur.id as purchase_id,
                pur.purchase_date
            FROM purchases pur
            JOIN packages pkg ON pur.package_id = pkg.id
            WHERE pur.user_id = {user_id}
            ORDER BY pur.purchase_date;
            """
            
            pkg_result = run_sql_query(pkg_query)
            packages = []
            if pkg_result:
                for pkg_line in pkg_result.split('\n'):
                    if not pkg_line.strip():
                        continue
                    pkg_parts = pkg_line.split('|')
                    if len(pkg_parts) >= 5:
                        packages.append({
                            'id': pkg_parts[0].strip(),
                            'name': pkg_parts[1].strip() if pkg_parts[1] else 'N/A',
                            'amount': float(pkg_parts[2].strip()) if pkg_parts[2] else 0.0,
                            'purchase_id': pkg_parts[3].strip(),
                            'purchase_date': pkg_parts[4].strip() if len(pkg_parts) > 4 else None
                        })
            
            # Create key for matching (email or display_id)
            key = email.lower() if email else (display_id.upper() if display_id else None)
            if not key:
                continue
            
            db_users[key] = {
                'id': user_id,
                'display_id': display_id,
                'email': email,
                'name': name,
                'packages': packages
            }
    
    print(f"✅ Found {len(db_users)} users in database")
    return db_users

def compare_users_and_packages():
    """Main comparison function"""
    print("=" * 100)
    print("🔍 COMPARING EXCEL vs LOCAL DATABASE")
    print("=" * 100)
    print()
    
    # Get data
    excel_users = get_excel_users()
    db_users = get_db_users_with_packages()
    
    print()
    print("=" * 100)
    print("📊 COMPARISON RESULTS")
    print("=" * 100)
    print()
    
    matches = []
    mismatches = []
    missing_in_db = []
    missing_in_excel = []
    
    # Check Excel users
    for key, excel_data in excel_users.items():
        db_data = db_users.get(key)
        
        if not db_data:
            missing_in_db.append({
                'key': key,
                'email': excel_data['email'],
                'member_id': excel_data['member_id'],
                'name': excel_data['name'],
                'excel_packages': len(excel_data['packages'])
            })
            continue
        
        excel_pkg_list = excel_data['packages']
        db_pkg_list = db_data['packages']
        
        # Match packages by amount
        matched_packages = []
        unmatched_excel = []
        unmatched_db = []
        
        for excel_pkg in excel_pkg_list:
            matched = False
            for db_pkg in db_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    matched_packages.append({
                        'excel': excel_pkg,
                        'db': db_pkg
                    })
                    break
            if not matched:
                unmatched_excel.append(excel_pkg)
        
        for db_pkg in db_pkg_list:
            matched = False
            for excel_pkg in excel_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    break
            if not matched:
                unmatched_db.append(db_pkg)
        
        # Check if everything matches
        if len(excel_pkg_list) == len(db_pkg_list) == len(matched_packages) and not unmatched_excel and not unmatched_db:
            matches.append({
                'key': key,
                'email': excel_data['email'],
                'member_id': excel_data['member_id'],
                'name': excel_data['name'] or db_data['name'],
                'excel_packages': len(excel_pkg_list),
                'db_packages': len(db_pkg_list)
            })
        else:
            mismatches.append({
                'key': key,
                'email': excel_data['email'],
                'member_id': excel_data['member_id'],
                'name': excel_data['name'] or db_data['name'],
                'excel_packages': excel_pkg_list,
                'db_packages': db_pkg_list,
                'matched': matched_packages,
                'unmatched_excel': unmatched_excel,
                'unmatched_db': unmatched_db
            })
    
    # Check DB users not in Excel
    for key, db_data in db_users.items():
        if key not in excel_users:
            missing_in_excel.append({
                'key': key,
                'email': db_data['email'],
                'display_id': db_data['display_id'],
                'name': db_data['name'],
                'db_packages': len(db_data['packages'])
            })
    
    # Print results
    print(f"✅ MATCHES: {len(matches)}")
    print(f"⚠️  MISMATCHES: {len(mismatches)}")
    print(f"❌ Missing in DB: {len(missing_in_db)}")
    print(f"❌ Missing in Excel: {len(missing_in_excel)}")
    print()
    
    # Show matches (first 10)
    if matches:
        print("=" * 100)
        print("✅ MATCHED USERS (showing first 10):")
        print("=" * 100)
        for i, match in enumerate(matches[:10], 1):
            print(f"{i}. {match['name'] or 'N/A'} ({match['email'] or match['member_id']})")
            print(f"   Excel Packages: {match['excel_packages']}, DB Packages: {match['db_packages']}")
        if len(matches) > 10:
            print(f"   ... and {len(matches) - 10} more matches")
        print()
    
    # Show mismatches
    if mismatches:
        print("=" * 100)
        print("⚠️  MISMATCHED USERS:")
        print("=" * 100)
        for i, mismatch in enumerate(mismatches, 1):
            print(f"\n{i}. {mismatch['name'] or 'N/A'} ({mismatch['email'] or mismatch['member_id']})")
            print(f"   Excel Packages ({len(mismatch['excel_packages'])}):")
            for pkg in mismatch['excel_packages']:
                print(f"      - {pkg['name']}: ₹{pkg['amount']:.2f}")
            print(f"   DB Packages ({len(mismatch['db_packages'])}):")
            for pkg in mismatch['db_packages']:
                print(f"      - {pkg['name']}: ₹{pkg['amount']:.2f}")
            if mismatch['unmatched_excel']:
                print(f"   ⚠️  Excel packages not in DB ({len(mismatch['unmatched_excel'])}):")
                for pkg in mismatch['unmatched_excel']:
                    print(f"      - {pkg['name']}: ₹{pkg['amount']:.2f}")
            if mismatch['unmatched_db']:
                print(f"   ⚠️  DB packages not in Excel ({len(mismatch['unmatched_db'])}):")
                for pkg in mismatch['unmatched_db']:
                    print(f"      - {pkg['name']}: ₹{pkg['amount']:.2f}")
        print()
    
    # Show missing in DB
    if missing_in_db:
        print("=" * 100)
        print(f"❌ USERS IN EXCEL BUT NOT IN DB ({len(missing_in_db)}):")
        print("=" * 100)
        for i, missing in enumerate(missing_in_db[:20], 1):
            print(f"{i}. {missing['name'] or 'N/A'} ({missing['email'] or missing['member_id']}) - {missing['excel_packages']} packages")
        if len(missing_in_db) > 20:
            print(f"   ... and {len(missing_in_db) - 20} more")
        print()
    
    # Show missing in Excel
    if missing_in_excel:
        print("=" * 100)
        print(f"❌ USERS IN DB BUT NOT IN EXCEL ({len(missing_in_excel)}):")
        print("=" * 100)
        for i, missing in enumerate(missing_in_excel[:20], 1):
            print(f"{i}. {missing['name'] or 'N/A'} ({missing['email'] or missing['display_id']}) - {missing['db_packages']} packages")
        if len(missing_in_excel) > 20:
            print(f"   ... and {len(missing_in_excel) - 20} more")
        print()
    
    # Summary
    print("=" * 100)
    print("📈 SUMMARY")
    print("=" * 100)
    print(f"Total Excel Users: {len(excel_users)}")
    print(f"Total DB Users: {len(db_users)}")
    print(f"Matched Users: {len(matches)}")
    print(f"Mismatched Users: {len(mismatches)}")
    print(f"Missing in DB: {len(missing_in_db)}")
    print(f"Missing in Excel: {len(missing_in_excel)}")
    print()
    
    return {
        'matches': len(matches),
        'mismatches': len(mismatches),
        'missing_in_db': len(missing_in_db),
        'missing_in_excel': len(missing_in_excel)
    }

if __name__ == '__main__':
    compare_users_and_packages()

