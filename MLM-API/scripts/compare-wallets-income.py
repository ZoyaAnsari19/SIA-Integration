#!/usr/bin/env python3
"""
Compare Excel vs Database:
- Spot Wallet
- Other Wallet  
- Income (Self + Global)
"""

import openpyxl
import subprocess
import json

# Configuration
EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        print(f"Query error: {result.stderr[:200]}")
        return None
    
    return result.stdout.strip()

def get_excel_wallets_income():
    """Read wallet and income data from Excel"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return {}
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    excel_data = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
        
        if not email and not member_id:
            continue
        
        # Get wallet amounts
        spot_amt = ws.cell(row_idx, headers.get('Spot amt.', 0)).value if headers.get('Spot amt.') else 0
        other_amt = ws.cell(row_idx, headers.get('Other amt.', 0)).value if headers.get('Other amt.') else 0
        
        spot_balance = float(spot_amt) if spot_amt and isinstance(spot_amt, (int, float)) else 0.0
        other_balance = float(other_amt) if other_amt and isinstance(other_amt, (int, float)) else 0.0
        
        # Get income (Self + Global) - sum of all package incomes
        total_income = 0.0
        for pack_num in range(1, 9):
            income_col = f'Pack. {pack_num} self + global'
            income_val = ws.cell(row_idx, headers.get(income_col, 0)).value if headers.get(income_col) else None
            if income_val:
                try:
                    total_income += float(income_val) if isinstance(income_val, (int, float)) else 0.0
                except:
                    pass
        
        # Create key for matching (email or member_id)
        key = email.lower() if email else (member_id.upper() if member_id else None)
        if not key:
            continue
        
        excel_data[key] = {
            'email': email,
            'member_id': member_id,
            'name': name,
            'spot_balance': spot_balance,
            'other_balance': other_balance,
            'total_income': total_income
        }
    
    print(f"✅ Found {len(excel_data)} users in Excel")
    return excel_data

def get_db_wallets_income():
    """Get wallet and income data from database"""
    print("🗄️  Reading database...")
    
    # Get all users with wallets and income
    query = """
    SELECT 
        u.id,
        u.display_id,
        u.email,
        u.name,
        COALESCE(ub.spot_balance, 0) as spot_balance,
        COALESCE(ub.other_balance, 0) as other_balance,
        COALESCE(income.self_income, 0) as self_income,
        COALESCE(income.global_income, 0) as global_income,
        COALESCE(income.self_income, 0) + COALESCE(income.global_income, 0) as total_income
    FROM users u
    LEFT JOIN user_balances ub ON ub.user_id = u.id
    LEFT JOIN (
        SELECT 
            receiver_user_id as user_id,
            SUM(CASE WHEN commission_type = 'SELF' THEN amount ELSE 0 END) as self_income,
            SUM(CASE WHEN commission_type = 'GLOBAL_HELPING' THEN amount ELSE 0 END) as global_income
        FROM ledger_entries
        GROUP BY receiver_user_id
    ) income ON income.user_id = u.id
    ORDER BY u.id;
    """
    
    result = run_sql_query(query)
    if not result:
        print("❌ Failed to get data from database")
        return {}
    
    db_data = {}
    
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 9:
            user_id = parts[0].strip()
            display_id = parts[1].strip() if parts[1] else None
            email = parts[2].strip() if parts[2] else None
            name = parts[3].strip() if parts[3] else None
            spot_balance = float(parts[4].strip()) if parts[4] else 0.0
            other_balance = float(parts[5].strip()) if parts[5] else 0.0
            self_income = float(parts[6].strip()) if parts[6] else 0.0
            global_income = float(parts[7].strip()) if parts[7] else 0.0
            total_income = float(parts[8].strip()) if parts[8] else 0.0
            
            # Create key for matching (email or display_id)
            key = email.lower() if email else (display_id.upper() if display_id else None)
            if not key:
                continue
            
            db_data[key] = {
                'id': user_id,
                'display_id': display_id,
                'email': email,
                'name': name,
                'spot_balance': spot_balance,
                'other_balance': other_balance,
                'self_income': self_income,
                'global_income': global_income,
                'total_income': total_income
            }
    
    print(f"✅ Found {len(db_data)} users in database")
    return db_data

def compare_wallets_income():
    """Main comparison function"""
    print("=" * 100)
    print("🔍 COMPARING WALLETS & INCOME: Excel vs Database")
    print("=" * 100)
    print()
    
    excel_data = get_excel_wallets_income()
    db_data = get_db_wallets_income()
    
    print()
    print("=" * 100)
    print("📊 COMPARISON RESULTS")
    print("=" * 100)
    print()
    
    spot_mismatches = []
    other_mismatches = []
    income_mismatches = []
    all_match = []
    missing_in_db = []
    missing_in_excel = []
    
    # Tolerance for floating point comparison
    TOLERANCE = 0.01
    
    # Check Excel users
    for key, excel_user in excel_data.items():
        db_user = db_data.get(key)
        
        if not db_user:
            missing_in_db.append(excel_user)
            continue
        
        # Compare Spot Wallet
        spot_diff = abs(excel_user['spot_balance'] - db_user['spot_balance'])
        spot_match = spot_diff < TOLERANCE
        
        # Compare Other Wallet
        other_diff = abs(excel_user['other_balance'] - db_user['other_balance'])
        other_match = other_diff < TOLERANCE
        
        # Compare Income (Self + Global)
        income_diff = abs(excel_user['total_income'] - db_user['total_income'])
        income_match = income_diff < TOLERANCE
        
        if spot_match and other_match and income_match:
            all_match.append({
                'key': key,
                'email': excel_user['email'] or db_user['email'],
                'member_id': excel_user['member_id'] or db_user['display_id'],
                'name': excel_user['name'] or db_user['name']
            })
        else:
            mismatch = {
                'key': key,
                'email': excel_user['email'] or db_user['email'],
                'member_id': excel_user['member_id'] or db_user['display_id'],
                'name': excel_user['name'] or db_user['name'],
                'user_id': db_user['id']
            }
            
            if not spot_match:
                mismatch['spot'] = {
                    'excel': excel_user['spot_balance'],
                    'db': db_user['spot_balance'],
                    'diff': excel_user['spot_balance'] - db_user['spot_balance']
                }
                spot_mismatches.append(mismatch)
            
            if not other_match:
                mismatch['other'] = {
                    'excel': excel_user['other_balance'],
                    'db': db_user['other_balance'],
                    'diff': excel_user['other_balance'] - db_user['other_balance']
                }
                other_mismatches.append(mismatch)
            
            if not income_match:
                mismatch['income'] = {
                    'excel': excel_user['total_income'],
                    'db': db_user['total_income'],
                    'db_self': db_user['self_income'],
                    'db_global': db_user['global_income'],
                    'diff': excel_user['total_income'] - db_user['total_income']
                }
                income_mismatches.append(mismatch)
    
    # Check DB users not in Excel
    for key, db_user in db_data.items():
        if key not in excel_data:
            missing_in_excel.append(db_user)
    
    # Print results
    print(f"✅ ALL MATCH: {len(all_match)} users")
    print(f"⚠️  Spot Wallet Mismatches: {len(spot_mismatches)} users")
    print(f"⚠️  Other Wallet Mismatches: {len(other_mismatches)} users")
    print(f"⚠️  Income Mismatches: {len(income_mismatches)} users")
    print(f"❌ Missing in DB: {len(missing_in_db)} users")
    print(f"❌ Missing in Excel: {len(missing_in_excel)} users")
    print()
    
    # Show all matched users
    if all_match:
        print("=" * 100)
        print(f"✅ ALL MATCHED USERS ({len(all_match)} users):")
        print("=" * 100)
        for i, match in enumerate(all_match, 1):
            print(f"{i}. {match['name'] or 'N/A'} ({match['email'] or match['member_id']})")
        print()
    
    # Show Spot Wallet mismatches
    if spot_mismatches:
        print("=" * 100)
        print(f"⚠️  SPOT WALLET MISMATCHES ({len(spot_mismatches)} users):")
        print("=" * 100)
        for i, mismatch in enumerate(spot_mismatches[:20], 1):
            spot_info = mismatch.get('spot', {})
            print(f"{i}. {mismatch['name'] or 'N/A'} ({mismatch['email'] or mismatch['member_id']})")
            print(f"   Excel: ₹{spot_info.get('excel', 0):.2f}")
            print(f"   DB: ₹{spot_info.get('db', 0):.2f}")
            print(f"   Difference: ₹{spot_info.get('diff', 0):.2f}")
        if len(spot_mismatches) > 20:
            print(f"   ... and {len(spot_mismatches) - 20} more")
        print()
    
    # Show Other Wallet mismatches
    if other_mismatches:
        print("=" * 100)
        print(f"⚠️  OTHER WALLET MISMATCHES ({len(other_mismatches)} users):")
        print("=" * 100)
        for i, mismatch in enumerate(other_mismatches[:20], 1):
            other_info = mismatch.get('other', {})
            print(f"{i}. {mismatch['name'] or 'N/A'} ({mismatch['email'] or mismatch['member_id']})")
            print(f"   Excel: ₹{other_info.get('excel', 0):.2f}")
            print(f"   DB: ₹{other_info.get('db', 0):.2f}")
            print(f"   Difference: ₹{other_info.get('diff', 0):.2f}")
        if len(other_mismatches) > 20:
            print(f"   ... and {len(other_mismatches) - 20} more")
        print()
    
    # Show Income mismatches
    if income_mismatches:
        print("=" * 100)
        print(f"⚠️  INCOME (SELF + GLOBAL) MISMATCHES ({len(income_mismatches)} users):")
        print("=" * 100)
        for i, mismatch in enumerate(income_mismatches[:20], 1):
            income_info = mismatch.get('income', {})
            print(f"{i}. {mismatch['name'] or 'N/A'} ({mismatch['email'] or mismatch['member_id']})")
            print(f"   Excel Total: ₹{income_info.get('excel', 0):.2f}")
            print(f"   DB Total: ₹{income_info.get('db', 0):.2f}")
            print(f"     - Self: ₹{income_info.get('db_self', 0):.2f}")
            print(f"     - Global: ₹{income_info.get('db_global', 0):.2f}")
            print(f"   Difference: ₹{income_info.get('diff', 0):.2f}")
        if len(income_mismatches) > 20:
            print(f"   ... and {len(income_mismatches) - 20} more")
        print()
    
    # Summary
    print("=" * 100)
    print("📈 SUMMARY")
    print("=" * 100)
    print(f"Total Excel Users: {len(excel_data)}")
    print(f"Total DB Users: {len(db_data)}")
    print(f"✅ All Match: {len(all_match)}")
    print(f"⚠️  Spot Wallet Mismatches: {len(spot_mismatches)}")
    print(f"⚠️  Other Wallet Mismatches: {len(other_mismatches)}")
    print(f"⚠️  Income Mismatches: {len(income_mismatches)}")
    print(f"❌ Missing in DB: {len(missing_in_db)}")
    print(f"❌ Missing in Excel: {len(missing_in_excel)}")
    print()
    
    return {
        'all_match': len(all_match),
        'spot_mismatches': len(spot_mismatches),
        'other_mismatches': len(other_mismatches),
        'income_mismatches': len(income_mismatches),
        'missing_in_db': len(missing_in_db),
        'missing_in_excel': len(missing_in_excel)
    }

if __name__ == '__main__':
    compare_wallets_income()

