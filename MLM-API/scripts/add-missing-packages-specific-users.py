#!/usr/bin/env python3
"""
Add missing packages for specific users based on Excel data
Includes global IDs and income from Excel
"""

import openpyxl
import subprocess
import json
import re
from datetime import datetime, timedelta

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

# Target users to fix
TARGET_USERS = ['SIA00748', 'SIA00299', 'SIA00602', 'SIA00446', 'SIA00405', 'SIA00601', 'SIA01115', 'SIA00209']

def run_sql_query(query, description="", return_id=False):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        if description:
            error_msg = result.stderr[:200] if result.stderr else "Unknown error"
            print(f"    ❌ Error: {error_msg}")
        return None
    
    output = result.stdout.strip()
    
    if return_id and output:
        lines = output.split('\n')
        for line in lines:
            line = line.strip()
            if line and not line.startswith('INSERT') and line.isdigit():
                return line
        numbers = re.findall(r'\d+', output)
        if numbers:
            return numbers[-1]
    
    return output

def find_package_by_amount(amount):
    """Find package_id by amount"""
    query = f"SELECT id, name, global_ids FROM packages WHERE price = {amount} LIMIT 1;"
    result = run_sql_query(query)
    if result:
        parts = result.split('|')
        if len(parts) >= 3:
            return {
                'id': int(parts[0].strip()),
                'name': parts[1].strip() if parts[1] else None,
                'global_ids': int(parts[2]) if parts[2] and parts[2].strip() != 'NULL' else None
            }
    return None

def get_user_id(display_id, email):
    """Get user ID from display_id or email"""
    if display_id:
        query = f"SELECT id FROM users WHERE display_id = '{display_id}' LIMIT 1;"
        result = run_sql_query(query)
        if result:
            return result.strip()
    
    if email:
        query = f"SELECT id FROM users WHERE LOWER(email) = LOWER('{email}') LIMIT 1;"
        result = run_sql_query(query)
        if result:
            return result.strip()
    
    return None

def get_existing_packages(user_id):
    """Get existing packages for user"""
    query = f"SELECT package_id, amount FROM purchases WHERE user_id = {user_id} AND status = 'completed';"
    result = run_sql_query(query)
    existing = {}
    if result:
        for line in result.split('\n'):
            if line.strip():
                parts = line.split('|')
                if len(parts) >= 2:
                    pkg_id = int(parts[0].strip())
                    amount = float(parts[1].strip())
                    if amount not in existing:
                        existing[amount] = []
                    existing[amount].append(pkg_id)
    return existing

def create_purchase(user_id, package_id, amount, income, used_global_id, package_global_ids):
    """Create purchase record with income and global IDs"""
    purchased_at = datetime.now()
    
    # Check if expired (2x amount)
    if income >= amount * 2:
        active_until = purchased_at  # Mark as expired
    else:
        active_until = purchased_at + timedelta(days=365)
    
    purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
    active_until_str = active_until.strftime('%Y-%m-%d %H:%M:%S')
    
    # Use Excel's used global id if available
    if used_global_id is not None and used_global_id > 0:
        global_ids_value = used_global_id
    else:
        global_ids_value = package_global_ids if package_global_ids else 'NULL'
    
    global_ids_sql = global_ids_value if global_ids_value != 'NULL' else 'NULL'
    
    query = f"""
    INSERT INTO purchases (
      user_id,
      package_id,
      purchase_type,
      amount,
      purchased_at,
      active_until,
      status,
      is_manual,
      is_renewal,
      income,
      effective_global_ids
    ) VALUES (
      {user_id},
      {package_id},
      'DIRECT_PACKAGE',
      {amount},
      '{purchased_at_str}',
      '{active_until_str}',
      'completed',
      true,
      false,
      {income},
      {global_ids_sql}
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, f"Creating purchase ₹{amount}", return_id=True)
    return result

print("=" * 100)
print("📦 ADD MISSING PACKAGES FOR SPECIFIC USERS")
print("=" * 100)
print()

# Read Excel
print("📄 Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

users_to_process = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
    
    if not member_id or str(member_id).upper() not in TARGET_USERS:
        continue
    
    packages = []
    for pack_num in range(1, 9):
        pack_amt_col = f'Pack. {pack_num} amt.'
        pack_name_col = f'Pack. {pack_num} name'
        pack_global_col = f'Pack. {pack_num} used global id'
        pack_income_col = f'Pack. {pack_num} self + global'
        
        pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
        if pack_amt and isinstance(pack_amt, (int, float)) and float(pack_amt) > 0:
            pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
            pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else 0
            pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else 0
            
            used_global = int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
            income = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0
            
            packages.append({
                'amount': float(pack_amt),
                'name': pack_name,
                'used_global_id': used_global,
                'income': income
            })
    
    if packages:
        users_to_process.append({
            'display_id': str(member_id).upper(),
            'email': email,
            'name': name,
            'packages': packages
        })

print(f"✅ Found {len(users_to_process)} users to process")
print()

total_added = 0

for user_data in users_to_process:
    print(f"\n👤 {user_data['name']} ({user_data['display_id']})")
    print("-" * 100)
    
    # Get user ID
    user_id = get_user_id(user_data['display_id'], user_data['email'])
    if not user_id:
        print(f"  ❌ User not found in DB")
        continue
    
    print(f"  User ID: {user_id}")
    
    # Get existing packages
    existing = get_existing_packages(user_id)
    print(f"  Existing packages: {list(existing.keys())}")
    
    # Find missing packages
    missing = []
    for pkg in user_data['packages']:
        amount = pkg['amount']
        if amount not in existing:
            missing.append(pkg)
        else:
            # Check if we need to add another instance (for duplicates)
            # For now, skip if exists
            pass
    
    if not missing:
        print(f"  ✅ All packages already exist")
        continue
    
    print(f"  Missing packages: {len(missing)}")
    
    # Add missing packages
    for pkg in missing:
        print(f"\n  📦 Package: {pkg['name']} - ₹{pkg['amount']:,.2f}")
        print(f"     Excel Global IDs: {pkg['used_global_id']}")
        print(f"     Excel Income: ₹{pkg['income']:,.2f}")
        
        # Find package_id
        package_info = find_package_by_amount(pkg['amount'])
        if not package_info:
            print(f"     ❌ Package not found for amount ₹{pkg['amount']:,.2f}")
            continue
        
        package_id = package_info['id']
        package_global_ids = package_info['global_ids'] or 0
        
        print(f"     Package ID: {package_id}, Total Global IDs: {package_global_ids}")
        
        # Create purchase
        purchase_id = create_purchase(
            user_id=user_id,
            package_id=package_id,
            amount=pkg['amount'],
            income=pkg['income'],
            used_global_id=pkg['used_global_id'],
            package_global_ids=package_global_ids
        )
        
        if purchase_id:
            print(f"     ✅ Purchase created: ID {purchase_id}")
            print(f"     ✅ Global IDs set to: {pkg['used_global_id']}")
            print(f"     ✅ Income set to: ₹{pkg['income']:,.2f}")
            total_added += 1
        else:
            print(f"     ❌ Failed to create purchase")

print()
print("=" * 100)
print(f"✅ COMPLETE: Added {total_added} missing packages")
print("=" * 100)

