#!/usr/bin/env python3
"""
Migration script for NEW users from products-export-3.xlsx

Process:
1. Check if user exists in DB → Skip if exists
2. Find referrer in DB → Error if not found
3. Create user account → Password: 123456
4. Create user tree paths → All ancestors
5. Create wallet → Excel amounts
6. IF user has packages:
   - Create purchase records
   - Process SPOT commissions
7. IF user has NO packages:
   - Skip purchase creation
"""

import openpyxl
import subprocess
import json
from datetime import datetime, timedelta

# Configuration
# Database connection matches MLM-API .env:
# DATABASE_URL=postgresql://mlm_user:mlm_password@localhost:5435/mlm_commission
# Container: mlm-prod-dump (port 5435 -> 5432)
EXCEL_FILE = 'products-export-3.xlsx'
USE_LOCAL_DB = True  # Set to True to use local Docker container, False for Kubernetes
LOCAL_CONTAINER = 'mlm-prod-dump'  # ✅ This is the database MLM-API is connected to
KUBECONFIG = 'azure-kube/0ffdcdf4-849b-4521-9868-be1000865e08'
NAMESPACE = 'mlm'
POD_NAME = 'postgres-0'
DB_USER = 'mlm_user'  # Matches MLM-API .env
DB_NAME = 'mlm_commission'  # Matches MLM-API .env

DEFAULT_PASSWORD = '123456'

def run_sql_query(query, description="", return_id=False):
    """Execute SQL query via docker exec (local) or kubectl exec (remote)"""
    if description:
        print(f"    {description}...")
    
    # Clean query (remove extra whitespace, newlines)
    query = ' '.join(query.split())
    
    if USE_LOCAL_DB:
        # Use local Docker container - no need for complex escaping
        # Just escape single quotes for shell: ' -> '\''
        escaped_query = query.replace("'", "'\\''")
        cmd = [
            'docker', 'exec', LOCAL_CONTAINER,
            'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
        ]
    else:
        # Use Kubernetes pod - escape for kubectl exec
        escaped_query = query.replace("'", "'\"'\"'")
        cmd = [
            'kubectl', '--kubeconfig', KUBECONFIG,
            'exec', '-n', NAMESPACE, POD_NAME, '--',
            'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
        ]
    
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=60
    )
    
    if result.returncode != 0:
        if description:
            error_msg = result.stderr[:200] if result.stderr else "Unknown error"
            print(f"    ❌ Error: {error_msg}")
        return None
    
    output = result.stdout.strip()
    
    # If RETURNING id, parse the result properly
    if return_id and output:
        # psql returns "INSERT 0 1\n<id>" for RETURNING queries
        lines = output.split('\n')
        for line in lines:
            line = line.strip()
            # Skip INSERT messages, get the actual ID
            if line and not line.startswith('INSERT') and line.isdigit():
                return line
        # Fallback: extract last number from output
        numbers = re.findall(r'\d+', output)
        if numbers:
            return numbers[-1]
    
    return output

def check_user_exists(email=None, display_id=None, user_id=None):
    """Check if user exists in database"""
    if email:
        escaped_email = email.replace("'", "''")
        query = f"SELECT id FROM users WHERE LOWER(email) = LOWER('{escaped_email}') LIMIT 1;"
        result = run_sql_query(query)
        if result:
            return True
    
    if display_id:
        escaped_display_id = display_id.replace("'", "''")
        query = f"SELECT id FROM users WHERE LOWER(display_id) = LOWER('{escaped_display_id}') LIMIT 1;"
        result = run_sql_query(query)
        if result:
            return True
    
    if user_id:
        query = f"SELECT id FROM users WHERE id = {user_id} LIMIT 1;"
        result = run_sql_query(query)
        if result:
            return True
    
    return False

def find_referrer_by_display_id(sponsor_id):
    """Find referrer user by display_id (Sponsor ID)"""
    if not sponsor_id:
        return None
    
    escaped_sponsor_id = sponsor_id.replace("'", "''")
    query = f"SELECT id, display_id, name, email FROM users WHERE LOWER(display_id) = LOWER('{escaped_sponsor_id}') LIMIT 1;"
    
    result = run_sql_query(query, f"Finding referrer: {sponsor_id}")
    if result:
        parts = result.split('|')
        if len(parts) >= 4:
            return {
                'id': parts[0].strip(),
                'display_id': parts[1].strip() if parts[1] else None,
                'name': parts[2].strip() if parts[2] else None,
                'email': parts[3].strip() if parts[3] else None
            }
    
    return None

def create_user(user_data, referrer_id):
    """Create new user account"""
    escaped_email = user_data['email'].replace("'", "''") if user_data['email'] else 'NULL'
    escaped_display_id = user_data['member_id'].replace("'", "''") if user_data['member_id'] else 'NULL'
    escaped_name = user_data['name'].replace("'", "''") if user_data['name'] else 'NULL'
    escaped_phone = user_data['phone'].replace("'", "''") if user_data['phone'] else 'NULL'
    
    referrer_sql = referrer_id if referrer_id else 'NULL'
    
    query = f"""
    INSERT INTO users (
      email,
      display_id,
      name,
      phone,
      password_hash,
      password_plain,
      referrer_user_id,
      role,
      status,
      created_at,
      updated_at
    ) VALUES (
      {'NULL' if escaped_email == 'NULL' else f"'{escaped_email}'"},
      {'NULL' if escaped_display_id == 'NULL' else f"'{escaped_display_id}'"},
      {'NULL' if escaped_name == 'NULL' else f"'{escaped_name}'"},
      {'NULL' if escaped_phone == 'NULL' else f"'{escaped_phone}'"},
      '{DEFAULT_PASSWORD}',
      '{DEFAULT_PASSWORD}',
      {referrer_sql},
      'STUDENT',
      'active',
      NOW(),
      NOW()
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, "Creating user account", return_id=True)
    return result

def create_user_tree_paths(user_id, referrer_id):
    """Create user tree paths for new user"""
    # 1. Self path (depth 0)
    query1 = f"INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth) VALUES ({user_id}, {user_id}, 0) ON CONFLICT DO NOTHING;"
    run_sql_query(query1, "Creating self path")
    
    if not referrer_id:
        return
    
    # 2. Direct referrer path (depth 1)
    query2 = f"INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth) VALUES ({referrer_id}, {user_id}, 1) ON CONFLICT DO NOTHING;"
    run_sql_query(query2, "Creating direct referrer path")
    
    # 3. Get all ancestors of referrer and create paths with +1 depth
    query3 = f"""
    INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth)
    SELECT ancestor_id, {user_id}, depth + 1
    FROM user_tree_paths
    WHERE descendant_id = {referrer_id}
      AND ancestor_id != {referrer_id}
    ON CONFLICT DO NOTHING;
    """
    run_sql_query(query3, "Creating ancestor paths")

def create_wallet(user_id, spot_balance, other_balance):
    """Create wallet for new user"""
    total_balance = spot_balance + other_balance
    
    query = f"""
    INSERT INTO user_balances (
      user_id,
      spot_balance,
      other_balance,
      balance
    ) VALUES (
      {user_id},
      {spot_balance},
      {other_balance},
      {total_balance}
    )
    ON CONFLICT (user_id) DO UPDATE SET
      spot_balance = {spot_balance},
      other_balance = {other_balance},
      balance = {total_balance};
    """
    
    run_sql_query(query, "Creating wallet")

def find_package_by_amount(amount):
    """Find package_id by amount"""
    query = f"SELECT id, name, global_ids FROM packages WHERE price = {amount} LIMIT 1;"
    
    result = run_sql_query(query, f"Finding package for amount: ₹{amount:,.2f}")
    if result:
        parts = result.split('|')
        if len(parts) >= 3:
            return {
                'id': int(parts[0].strip()),
                'name': parts[1].strip() if parts[1] else None,
                'global_ids': int(parts[2]) if parts[2] and parts[2].strip() != 'NULL' else None
            }
    
    return None

def check_if_reinvestment(user_id, purchase_date):
    """Check if this is a reinvestment"""
    query = f"SELECT COUNT(*) FROM purchases WHERE user_id = {user_id} AND status = 'completed' AND purchased_at < '{purchase_date}' AND active_until >= CURRENT_DATE;"
    
    result = run_sql_query(query)
    if result:
        try:
            count = int(result.strip())
            if count > 0:
                # Check if any previous purchase has not reached 2x
                query_2x = f"SELECT id, amount, income FROM purchases WHERE user_id = {user_id} AND status = 'completed' AND purchased_at < '{purchase_date}' AND active_until >= CURRENT_DATE;"
                result_2x = run_sql_query(query_2x)
                if result_2x:
                    for line in result_2x.split('\n'):
                        if line.strip():
                            parts = line.split('|')
                            if len(parts) >= 3:
                                try:
                                    amount = float(parts[1]) if parts[1] else 0
                                    income = float(parts[2]) if parts[2] else 0
                                    if income < amount * 2:
                                        return True
                                except:
                                    pass
        except:
            pass
    
    return False

def check_eligibility(user_id, level):
    """Check if user is eligible for a specific level"""
    query = f"SELECT eligibility::text FROM level_eligibility WHERE user_id = {user_id};"
    
    result = run_sql_query(query)
    if result:
        try:
            eligibility = json.loads(result)
            return bool(eligibility.get(str(level), False))
        except:
            try:
                cleaned = result.strip().strip('"').replace("\\\"", '"')
                eligibility = json.loads(cleaned)
                return bool(eligibility.get(str(level), False))
            except:
                pass
    
    return False

def get_spot_percent(level):
    """Get SPOT commission percentage for a level"""
    query = f"SELECT spot_commission_percent FROM levels WHERE level = {level};"
    
    result = run_sql_query(query)
    if result:
        try:
            percent = float(result.strip())
            return percent
        except:
            pass
    
    # Fallback to commission_rules
    query_fallback = f"SELECT percent FROM commission_rules WHERE type = 'LEVEL_SPOT' AND level = {level} LIMIT 1;"
    result = run_sql_query(query_fallback)
    if result:
        try:
            percent = float(result.strip())
            return percent
        except:
            pass
    
    return 0.0

def get_uplines(user_id, max_depth=10):
    """Get upline chain for a user"""
    query = f"SELECT ancestor_id, depth FROM user_tree_paths WHERE descendant_id = {user_id} AND depth >= 1 AND depth <= {max_depth} ORDER BY depth ASC;"
    
    result = run_sql_query(query)
    uplines = []
    
    if result:
        for line in result.split('\n'):
            if line.strip():
                parts = line.split('|')
                if len(parts) >= 2:
                    uplines.append({
                        'ancestor_id': parts[0].strip(),
                        'depth': int(parts[1].strip())
                    })
    
    return uplines

def create_purchase(user_id, package_id, amount, income, used_global_id, package_global_ids):
    """Create purchase record"""
    purchased_at = datetime.now()
    active_until = purchased_at + timedelta(days=365)  # Future date (display only)
    
    # Check if expired (2x amount)
    if income >= amount * 2:
        active_until = purchased_at  # Mark as expired
    
    purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
    active_until_str = active_until.strftime('%Y-%m-%d %H:%M:%S')
    
    # Use Excel's used global id if available, otherwise use package global_ids
    if used_global_id is not None:
        global_ids_value = used_global_id
    else:
        global_ids_value = package_global_ids if package_global_ids else 'NULL'
    
    global_ids_sql = global_ids_value if global_ids_value != 'NULL' else 'NULL'
    
    query = f"""
    INSERT INTO purchases (
      user_id,
      package_id,
      purchase_type,
      amount,
      purchased_at,
      active_until,
      status,
      is_manual,
      is_renewal,
      income,
      effective_global_ids
    ) VALUES (
      {user_id},
      {package_id},
      'DIRECT_PACKAGE',
      {amount},
      '{purchased_at_str}',
      '{active_until_str}',
      'completed',
      true,
      false,
      {income},
      {global_ids_sql}
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, "Creating purchase record", return_id=True)
    return result

def process_spot_commissions(buyer_id, purchase_id, purchase_amount, is_reinvestment):
    """Process SPOT commissions for uplines"""
    uplines = get_uplines(buyer_id, max_depth=10)
    
    pending_count = 0
    skipped_qualified = 0
    skipped_direct = 0
    
    for upline in uplines:
        depth = upline['depth']
        level = depth - 1  # depth 2 → level 1, etc.
        upline_id = upline['ancestor_id']
        
        # Skip direct referrer (Level 0, depth 1)
        # Direct referrer SPOT not processed in migration
        if depth == 1:
            skipped_direct += 1
            continue
        
        # For Team Uplines (Level 1-9, depth 2-10)
        if depth >= 2 and depth <= 10:
            # Get SPOT percentage
            spot_percent = get_spot_percent(level)
            if spot_percent == 0:
                continue
            
            # Calculate SPOT amount based on invested amount (purchase_amount from Excel)
            # spot_amount = invested_amount × level_spot_percent / 100
            spot_amount = purchase_amount * spot_percent / 100
            
            # Apply 50% reduction for reinvestments (Level 1+)
            if is_reinvestment and level >= 1:
                spot_amount = spot_amount * 0.5
            
            # Check eligibility: Is this upline qualified for this level?
            eligible = check_eligibility(upline_id, level)
            
            if eligible:
                # ✅ QUALIFIED UPLINE: Skip completely
                # Rule: "jo user inke level ko qualified karege unko naaa spot dena hai naaa pending mei add krna hai"
                # - NO SPOT commission
                # - NO pending_commissions entry
                # - Qualified users don't get retroactive commission in migration
                skipped_qualified += 1
                continue
            else:
                # ✅ UNQUALIFIED UPLINE: Add SPOT to pending_commissions
                # Rule: "jo upline user inke level ko qualified abhi nhi karte unke liye level k hisab se invested amount ka spot amount pending mei add karna hoga"
                # - Calculate: spot_amount = invested_amount × level_spot_percent / 100
                # - Add to pending_commissions table
                # - Will be released when upline qualifies for that level
                pending_query = f"""
                INSERT INTO pending_commissions (
                  receiver_user_id,
                  source_user_id,
                  purchase_id,
                  level,
                  commission_type,
                  amount,
                  metadata
                ) VALUES (
                  {upline_id},
                  {buyer_id},
                  {purchase_id},
                  {level},
                  'SPOT',
                  {spot_amount},
                  '{{"level": {level}, "depth": {depth}, "reason": "eligibility", "is_reinvestment": {str(is_reinvestment).lower()}}}'::jsonb
                )
                ON CONFLICT DO NOTHING;
                """
                
                result = run_sql_query(pending_query)
                if result is not None:
                    pending_count += 1
    
    return {
        'pending_count': pending_count,
        'skipped_qualified': skipped_qualified,
        'skipped_direct': skipped_direct
    }

def main():
    import os
    
    print("=" * 120)
    print("🚀 MIGRATION SCRIPT - NEW USERS FROM products-export-3.xlsx")
    print("=" * 120)
    print()
    
    # Step 1: Read Excel file
    print("📄 Step 1: Reading Excel file...")
    
    # Get absolute path to Excel file (from MLM root directory)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    mlm_root = os.path.dirname(os.path.dirname(script_dir))
    excel_path = os.path.join(mlm_root, EXCEL_FILE)
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        print(f"   Current directory: {os.getcwd()}")
        print(f"   Looking for: {EXCEL_FILE}")
        return
    
    print(f"   Reading from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    # Extract all users from Excel
    excel_users = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data[header] = cell_value
        
        user_id = str(row_data.get('ID', '')) if row_data.get('ID') else None
        member_id = str(row_data.get('Member ID', '')) if row_data.get('Member ID') else None
        user_name = str(row_data.get('Name', '')) if row_data.get('Name') else None
        user_email = str(row_data.get('Email', '')) if row_data.get('Email') else None
        user_phone = str(row_data.get('Phone', '')) if row_data.get('Phone') else None
        sponsor_id = str(row_data.get('Sponsor ID', '')) if row_data.get('Sponsor ID') else None
        sponsor_name = str(row_data.get('Sponsor Name', '')) if row_data.get('Sponsor Name') else None
        
        # Get wallet amounts
        spot_amt = row_data.get('Spot amt.')
        other_amt = row_data.get('Other amt.')
        excel_spot = float(spot_amt) if spot_amt is not None else 0
        excel_other = float(other_amt) if other_amt is not None else 0
        
        # Get packages
        packages = []
        for i in range(1, 9):
            pkg_amt_col = f'Pack. {i} amt.'
            pkg_name_col = f'Pack. {i} name'
            pkg_income_col = f'Pack. {i} self + global'
            pkg_used_global_id_col = f'Pack. {i} used global id'
            
            pkg_amount = row_data.get(pkg_amt_col)
            pkg_name = row_data.get(pkg_name_col)
            pkg_income = row_data.get(pkg_income_col)
            pkg_used_global_id = row_data.get(pkg_used_global_id_col)
            
            if pkg_amount is not None:
                try:
                    amount_float = float(pkg_amount)
                    income_float = float(pkg_income) if pkg_income is not None else 0
                    used_global_id = int(pkg_used_global_id) if pkg_used_global_id is not None and str(pkg_used_global_id) != 'N/A' else None
                    
                    packages.append({
                        'package_number': i,
                        'amount': amount_float,
                        'name': str(pkg_name) if pkg_name else 'N/A',
                        'income': income_float,
                        'used_global_id': used_global_id
                    })
                except:
                    pass
        
        if user_email or member_id or user_id:
            excel_users.append({
                'user_id': user_id,
                'member_id': member_id,
                'name': user_name,
                'email': user_email,
                'phone': user_phone,
                'sponsor_id': sponsor_id,
                'sponsor_name': sponsor_name,
                'excel_spot': excel_spot,
                'excel_other': excel_other,
                'packages': packages
            })
    
    print(f"📄 Total users in Excel: {len(excel_users)}")
    print()
    
    # Step 2: Filter new users (not in DB)
    print("🔍 Step 2: Filtering new users (not in database)...")
    new_users = []
    existing_users = []
    
    for excel_user in excel_users:
        exists = check_user_exists(
            email=excel_user['email'],
            display_id=excel_user['member_id'],
            user_id=excel_user['user_id']
        )
        
        if exists:
            existing_users.append(excel_user)
        else:
            new_users.append(excel_user)
    
    print(f"✅ Existing users in DB: {len(existing_users)}")
    print(f"🆕 New users to migrate: {len(new_users)}")
    print()
    
    if not new_users:
        print("✅ No new users to migrate!")
        return
    
    # Migration summary
    migration_summary = {
        'total_new_users': len(new_users),
        'users_created': 0,
        'users_skipped': 0,
        'wallets_created': 0,
        'purchases_created': 0,
        'pending_commissions_created': 0,
        'errors': []
    }
    
    # Step 3: Process each new user
    print("=" * 120)
    print("📦 Step 3: Processing New Users...")
    print("=" * 120)
    print()
    
    for idx, new_user in enumerate(new_users, 1):
        print(f"\n[{idx}/{len(new_users)}] Processing: {new_user['name']} ({new_user['email']})")
        
        # Step 1: Check if user exists (already checked, but double-check)
        if check_user_exists(email=new_user['email'], display_id=new_user['member_id']):
            print(f"  ⚠️  User already exists, skipping...")
            migration_summary['users_skipped'] += 1
            continue
        
        # Step 2: Find referrer
        referrer = find_referrer_by_display_id(new_user['sponsor_id'])
        if not referrer:
            error_msg = f"Referrer not found: {new_user['sponsor_id']}"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': new_user['name'],
                'error': error_msg
            })
            continue
        
        print(f"  ✅ Referrer found: {referrer['name']} (ID: {referrer['id']})")
        
        # Step 3: Create user account
        user_id = create_user(new_user, referrer['id'])
        if not user_id:
            error_msg = "Failed to create user account"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': new_user['name'],
                'error': error_msg
            })
            continue
        
        print(f"  ✅ User created: ID {user_id}")
        migration_summary['users_created'] += 1
        
        # Step 4: Create user tree paths
        print(f"  🌳 Creating user tree paths...")
        create_user_tree_paths(user_id, referrer['id'])
        print(f"  ✅ User tree paths created")
        
        # Step 5: Create wallet
        print(f"  💰 Creating wallet...")
        create_wallet(user_id, new_user['excel_spot'], new_user['excel_other'])
        print(f"  ✅ Wallet created: Spot ₹{new_user['excel_spot']:,.2f}, Other ₹{new_user['excel_other']:,.2f}")
        migration_summary['wallets_created'] += 1
        
        # Step 6: IF user has packages
        if new_user['packages']:
            print(f"  📦 Processing {len(new_user['packages'])} packages...")
            
            purchased_at = datetime.now()
            purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
            is_reinvestment = check_if_reinvestment(user_id, purchased_at_str)
            
            if is_reinvestment:
                print(f"  🔄 Reinvestment detected (50% reduction for Level 1+)")
            else:
                print(f"  ✨ First purchase (100% SPOT)")
            
            for package in new_user['packages']:
                # Map package
                package_info = find_package_by_amount(package['amount'])
                if not package_info:
                    error_msg = f"Package not found for amount: ₹{package['amount']:,.2f}"
                    print(f"  ❌ {error_msg}")
                    migration_summary['errors'].append({
                        'user': new_user['name'],
                        'error': error_msg
                    })
                    continue
                
                print(f"  ✅ Package mapped: {package_info['name']} (ID: {package_info['id']})")
                
                # Create purchase
                purchase_id = create_purchase(
                    user_id=user_id,
                    package_id=package_info['id'],
                    amount=package['amount'],
                    income=package['income'],
                    used_global_id=package['used_global_id'],
                    package_global_ids=package_info['global_ids']
                )
                
                if not purchase_id:
                    error_msg = "Failed to create purchase"
                    print(f"  ❌ {error_msg}")
                    continue
                
                print(f"  ✅ Purchase created: ID {purchase_id}")
                migration_summary['purchases_created'] += 1
                
                # Check expiry
                if package['income'] >= package['amount'] * 2:
                    print(f"  ⚠️  Package EXPIRED (income >= 2x amount)")
                else:
                    print(f"  ✅ Package ACTIVE (income < 2x amount)")
                
                # Process SPOT commissions
                print(f"  💵 Processing SPOT commissions...")
                spot_result = process_spot_commissions(
                    buyer_id=user_id,
                    purchase_id=purchase_id,
                    purchase_amount=package['amount'],
                    is_reinvestment=is_reinvestment
                )
                
                print(f"  ✅ SPOT processing complete:")
                print(f"     - Direct referrers skipped: {spot_result['skipped_direct']}")
                print(f"     - Qualified uplines skipped: {spot_result['skipped_qualified']}")
                print(f"     - Pending commissions created: {spot_result['pending_count']}")
                migration_summary['pending_commissions_created'] += spot_result['pending_count']
        else:
            print(f"  ⚠️  No packages found, skipping purchase creation")
    
    # Final summary
    print()
    print("=" * 120)
    print("📊 MIGRATION SUMMARY")
    print("=" * 120)
    print(f"Total New Users: {migration_summary['total_new_users']}")
    print(f"✅ Users Created: {migration_summary['users_created']}")
    print(f"⚠️  Users Skipped: {migration_summary['users_skipped']}")
    print(f"✅ Wallets Created: {migration_summary['wallets_created']}")
    print(f"✅ Purchases Created: {migration_summary['purchases_created']}")
    print(f"✅ Pending Commissions Created: {migration_summary['pending_commissions_created']}")
    print(f"❌ Errors: {len(migration_summary['errors'])}")
    
    if migration_summary['errors']:
        print("\n❌ ERRORS:")
        for error in migration_summary['errors']:
            print(f"  - {error['user']}: {error['error']}")
    
    # Save summary to file
    with open('new-users-migration-summary.json', 'w') as f:
        json.dump(migration_summary, f, indent=2, default=str)
    
    print("\n✅ Migration summary saved to new-users-migration-summary.json")
    print("=" * 120)

if __name__ == '__main__':
    main()

