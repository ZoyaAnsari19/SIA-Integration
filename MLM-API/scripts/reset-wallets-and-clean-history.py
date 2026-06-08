#!/usr/bin/env python3
"""
1. Reset all users' spot and other wallets to 0
2. Add Excel wallet amounts to all users
3. Clean all history tables (withdrawal, ledger, p2p, kyc, activation)
"""

import openpyxl
import subprocess

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query, description=""):
    """Execute SQL query"""
    if description:
        print(f"  {description}...")
    
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    
    if result.returncode == 0:
        return True, result.stdout
    else:
        return False, result.stderr

print("=" * 100)
print("🔄 RESET WALLETS & CLEAN HISTORY")
print("=" * 100)
print()

# Step 1: Reset all wallets to 0
print("📊 Step 1: Resetting all wallets to 0...")
success, output = run_sql_query("""
    UPDATE user_balances 
    SET spot_balance = 0,
        other_balance = 0,
        balance = 0;
""", "Resetting all wallets")

if success:
    print("  ✅ All wallets reset to 0")
else:
    print(f"  ❌ Error: {output[:200]}")
    exit(1)

print()

# Step 2: Clean history tables
print("🧹 Step 2: Cleaning history tables...")

history_tables = [
    ('ledger_entries', 'Ledger entries'),
    ('withdraw_requests', 'Withdrawal requests'),
    ('kyc_documents', 'KYC documents'),
    ('wallet_transactions', 'Wallet transactions'),
    ('pending_commissions', 'Pending commissions'),
    ('scheduled_commissions', 'Scheduled commissions'),
]

for table_name, description in history_tables:
    success, output = run_sql_query(f"DELETE FROM {table_name};", f"Cleaning {description}")
    if success:
        print(f"  ✅ {description} cleaned")
    else:
        # Table might not exist, that's okay
        if 'does not exist' not in output.lower():
            print(f"  ⚠️  {description}: {output[:100]}")

print()

# Step 3: Read Excel and update wallets
print("📄 Step 3: Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

users_data = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
    
    if not email and not member_id:
        continue
    
    spot_amt = ws.cell(row_idx, headers.get('Spot amt.', 0)).value if headers.get('Spot amt.') else 0
    other_amt = ws.cell(row_idx, headers.get('Other amt.', 0)).value if headers.get('Other amt.') else 0
    
    spot = float(spot_amt) if spot_amt and isinstance(spot_amt, (int, float)) else 0
    other = float(other_amt) if other_amt and isinstance(other_amt, (int, float)) else 0
    
    users_data.append({
        'email': email,
        'display_id': member_id,
        'name': name,
        'spot': spot,
        'other': other
    })

print(f"  ✅ Found {len(users_data)} users in Excel")
print()

# Step 4: Update wallets from Excel
print("💰 Step 4: Updating wallets from Excel...")

def get_user_id(display_id, email):
    """Get user ID"""
    if display_id:
        query = f"SELECT id FROM users WHERE display_id = '{display_id}' LIMIT 1;"
        escaped_query = query.replace("'", "'\\''")
        cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -c '{escaped_query}'"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    
    if email:
        query = f"SELECT id FROM users WHERE LOWER(email) = LOWER('{email}') LIMIT 1;"
        escaped_query = query.replace("'", "'\\''")
        cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -c '{escaped_query}'"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    
    return None

updated = 0
not_found = 0

for user_data in users_data:
    user_id = get_user_id(user_data['display_id'], user_data['email'])
    
    if not user_id:
        not_found += 1
        continue
    
    total_balance = user_data['spot'] + user_data['other']
    
    query = f"""
    INSERT INTO user_balances (
      user_id,
      spot_balance,
      other_balance,
      balance
    ) VALUES (
      {user_id},
      {user_data['spot']},
      {user_data['other']},
      {total_balance}
    )
    ON CONFLICT (user_id) DO UPDATE SET
      spot_balance = {user_data['spot']},
      other_balance = {user_data['other']},
      balance = {total_balance};
    """
    
    success, _ = run_sql_query(query)
    if success:
        updated += 1
        if updated <= 20:
            print(f"  ✅ {user_data['name'] or user_data['display_id'] or user_data['email']}: Spot ₹{user_data['spot']:,.2f}, Other ₹{user_data['other']:,.2f}")

print()
print("=" * 100)
print("✅ COMPLETE:")
print(f"   - Wallets reset: All users")
print(f"   - History tables cleaned: {len(history_tables)} tables")
print(f"   - Wallets updated from Excel: {updated} users")
print(f"   - Users not found in DB: {not_found}")
print("=" * 100)

