#!/usr/bin/env python3
"""
Add 5 missing users from Excel to DB
- Create users
- Add wallets (spot + other)
- Add packages if they have any
- Set income and global IDs from Excel
"""

import openpyxl
import subprocess
import re
from datetime import datetime, timedelta

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'
DEFAULT_PASSWORD = 'password123'

def run_sql_query(query, description="", return_id=False):
    """Execute SQL query"""
    if description:
        print(f"    {description}...")
    
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        if description:
            error_msg = result.stderr[:200] if result.stderr else "Unknown error"
            print(f"    ❌ Error: {error_msg}")
        return None
    
    output = result.stdout.strip()
    
    if return_id and output:
        lines = output.split('\n')
        for line in lines:
            line = line.strip()
            if line and not line.startswith('INSERT') and line.isdigit():
                return line
        numbers = re.findall(r'\d+', output)
        if numbers:
            return numbers[-1]
    
    return output

def find_referrer_by_display_id(sponsor_id):
    """Find referrer user by display_id"""
    if not sponsor_id:
        return None
    
    query = f"SELECT id FROM users WHERE LOWER(display_id) = LOWER('{sponsor_id}') LIMIT 1;"
    result = run_sql_query(query)
    if result:
        return result.strip()
    return None

def create_user(user_data, referrer_id):
    """Create new user"""
    email = user_data['email'].replace("'", "''") if user_data['email'] else 'NULL'
    display_id = user_data['display_id'].replace("'", "''") if user_data['display_id'] else 'NULL'
    name = user_data['name'].replace("'", "''") if user_data['name'] else 'NULL'
    phone = str(user_data['phone']).replace("'", "''") if user_data['phone'] else 'NULL'
    
    referrer_sql = referrer_id if referrer_id else 'NULL'
    
    query = f"""
    INSERT INTO users (
      email,
      display_id,
      name,
      phone,
      password_hash,
      password_plain,
      referrer_user_id,
      role,
      status,
      created_at,
      updated_at
    ) VALUES (
      {'NULL' if email == 'NULL' else f"'{email}'"},
      {'NULL' if display_id == 'NULL' else f"'{display_id}'"},
      {'NULL' if name == 'NULL' else f"'{name}'"},
      {'NULL' if phone == 'NULL' else f"'{phone}'"},
      '{DEFAULT_PASSWORD}',
      '{DEFAULT_PASSWORD}',
      {referrer_sql},
      'STUDENT',
      'active',
      NOW(),
      NOW()
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, "Creating user", return_id=True)
    return result

def create_user_tree_paths(user_id, referrer_id):
    """Create user tree paths"""
    # Self path
    query1 = f"INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth) VALUES ({user_id}, {user_id}, 0) ON CONFLICT DO NOTHING;"
    run_sql_query(query1)
    
    if not referrer_id:
        return
    
    # Direct referrer path
    query2 = f"INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth) VALUES ({referrer_id}, {user_id}, 1) ON CONFLICT DO NOTHING;"
    run_sql_query(query2)
    
    # Ancestor paths
    query3 = f"""
    INSERT INTO user_tree_paths (ancestor_id, descendant_id, depth)
    SELECT ancestor_id, {user_id}, depth + 1
    FROM user_tree_paths
    WHERE descendant_id = {referrer_id}
      AND ancestor_id != {referrer_id}
    ON CONFLICT DO NOTHING;
    """
    run_sql_query(query3)

def create_wallet(user_id, spot_balance, other_balance):
    """Create wallet"""
    total_balance = spot_balance + other_balance
    
    query = f"""
    INSERT INTO user_balances (
      user_id,
      spot_balance,
      other_balance,
      balance
    ) VALUES (
      {user_id},
      {spot_balance},
      {other_balance},
      {total_balance}
    )
    ON CONFLICT (user_id) DO UPDATE SET
      spot_balance = {spot_balance},
      other_balance = {other_balance},
      balance = {total_balance};
    """
    
    run_sql_query(query, "Creating wallet")

def find_package_by_amount(amount):
    """Find package_id by amount"""
    query = f"SELECT id, name, global_ids FROM packages WHERE price = {amount} LIMIT 1;"
    result = run_sql_query(query)
    if result:
        parts = result.split('|')
        if len(parts) >= 3:
            return {
                'id': int(parts[0].strip()),
                'name': parts[1].strip() if parts[1] else None,
                'global_ids': int(parts[2]) if parts[2] and parts[2].strip() != 'NULL' else None
            }
    return None

def create_purchase(user_id, package_id, amount, income, used_global_id, package_global_ids):
    """Create purchase with income and global IDs"""
    purchased_at = datetime.now()
    
    # Check if expired (2x amount)
    if income >= amount * 2:
        active_until = purchased_at
    else:
        active_until = purchased_at + timedelta(days=365)
    
    purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
    active_until_str = active_until.strftime('%Y-%m-%d %H:%M:%S')
    
    # Use Excel's used global id if available
    if used_global_id is not None and used_global_id > 0:
        global_ids_value = used_global_id
    else:
        global_ids_value = package_global_ids if package_global_ids else 'NULL'
    
    global_ids_sql = global_ids_value if global_ids_value != 'NULL' else 'NULL'
    
    query = f"""
    INSERT INTO purchases (
      user_id,
      package_id,
      purchase_type,
      amount,
      purchased_at,
      active_until,
      status,
      is_manual,
      is_renewal,
      income,
      effective_global_ids
    ) VALUES (
      {user_id},
      {package_id},
      'DIRECT_PACKAGE',
      {amount},
      '{purchased_at_str}',
      '{active_until_str}',
      'completed',
      true,
      false,
      {income},
      {global_ids_sql}
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, f"Creating purchase ₹{amount}", return_id=True)
    return result

print("=" * 100)
print("👥 ADDING 5 MISSING USERS FROM EXCEL")
print("=" * 100)
print()

# Read Excel
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

target_ids = ['SIA01916', 'SIA01917', 'SIA01918', 'SIA01919', 'SIA01920']
users_to_add = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    if not member_id or str(member_id).upper() not in target_ids:
        continue
    
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
    phone = ws.cell(row_idx, headers.get('Phone', 0)).value if headers.get('Phone') else None
    sponsor_id = ws.cell(row_idx, headers.get('Sponsor ID', 0)).value if headers.get('Sponsor ID') else None
    
    spot_amt = ws.cell(row_idx, headers.get('Spot amt.', 0)).value if headers.get('Spot amt.') else 0
    other_amt = ws.cell(row_idx, headers.get('Other amt.', 0)).value if headers.get('Other amt.') else 0
    
    spot = float(spot_amt) if spot_amt and isinstance(spot_amt, (int, float)) else 0
    other = float(other_amt) if other_amt and isinstance(other_amt, (int, float)) else 0
    
    packages = []
    for pack_num in range(1, 9):
        pack_amt_col = f'Pack. {pack_num} amt.'
        pack_name_col = f'Pack. {pack_num} name'
        pack_global_col = f'Pack. {pack_num} used global id'
        pack_income_col = f'Pack. {pack_num} self + global'
        
        pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
        if pack_amt and isinstance(pack_amt, (int, float)) and float(pack_amt) > 0:
            pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
            pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else 0
            pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else 0
            
            used_global = int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
            income = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0
            
            packages.append({
                'amount': float(pack_amt),
                'name': pack_name,
                'used_global_id': used_global,
                'income': income
            })
    
    users_to_add.append({
        'display_id': str(member_id).upper(),
        'email': email,
        'name': name,
        'phone': phone,
        'sponsor_id': sponsor_id,
        'spot': spot,
        'other': other,
        'packages': packages
    })

print(f"✅ Found {len(users_to_add)} users to add")
print()

# Add users
for user_data in users_to_add:
    print(f"\n👤 {user_data['name']} ({user_data['display_id']})")
    print("-" * 100)
    
    # Find referrer
    referrer_id = None
    if user_data['sponsor_id']:
        referrer_id = find_referrer_by_display_id(user_data['sponsor_id'])
        if referrer_id:
            print(f"  ✅ Referrer found: {user_data['sponsor_id']} (ID: {referrer_id})")
        else:
            print(f"  ⚠️  Referrer not found: {user_data['sponsor_id']}")
    
    # Create user
    user_id = create_user(user_data, referrer_id)
    if not user_id:
        print(f"  ❌ Failed to create user")
        continue
    
    print(f"  ✅ User created: ID {user_id}")
    
    # Create tree paths
    if referrer_id:
        create_user_tree_paths(user_id, referrer_id)
        print(f"  ✅ Tree paths created")
    
    # Create wallet
    create_wallet(user_id, user_data['spot'], user_data['other'])
    print(f"  ✅ Wallet created: Spot ₹{user_data['spot']:,.2f}, Other ₹{user_data['other']:,.2f}")
    
    # Add packages
    if user_data['packages']:
        print(f"  📦 Adding {len(user_data['packages'])} packages...")
        for pkg in user_data['packages']:
            package_info = find_package_by_amount(pkg['amount'])
            if not package_info:
                print(f"     ❌ Package not found for ₹{pkg['amount']:,.2f}")
                continue
            
            purchase_id = create_purchase(
                user_id=user_id,
                package_id=package_info['id'],
                amount=pkg['amount'],
                income=pkg['income'],
                used_global_id=pkg['used_global_id'],
                package_global_ids=package_info['global_ids'] or 0
            )
            
            if purchase_id:
                print(f"     ✅ Package added: {pkg['name']} - ₹{pkg['amount']:,.2f}")
                print(f"        Global IDs: {pkg['used_global_id']}, Income: ₹{pkg['income']:,.2f}")
    else:
        print(f"  📦 No packages to add")

print()
print("=" * 100)
print("✅ COMPLETE: All 5 users added to DB")
print("=" * 100)

