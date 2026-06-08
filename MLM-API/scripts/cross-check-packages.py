#!/usr/bin/env python3
"""
Cross-check packages for migrated new users
Compares Excel data with Database data
"""

import openpyxl
import subprocess
import json
from datetime import datetime

# Configuration
EXCEL_FILE = 'products-export-3.xlsx'
USE_LOCAL_DB = True
LOCAL_CONTAINER = 'mlm-postgres-local'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

# New user IDs range
NEW_USER_MIN_ID = 1956
NEW_USER_MAX_ID = 1982

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    
    if result.returncode != 0:
        return None
    
    return result.stdout.strip()

def get_excel_packages():
    """Read packages from Excel for new users"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return {}
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    excel_packages = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        user_id = ws.cell(row_idx, headers.get('ID', 0)).value if headers.get('ID') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        
        if not email and not member_id:
            continue
        
        # Get packages
        packages = []
        for pack_num in range(1, 9):
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_name_col = f'Pack. {pack_num} name'
            pack_income_col = f'Pack. {pack_num} self + global'
            pack_global_col = f'Pack. {pack_num} used global id'
            
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else None
            pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else None
            pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else None
            
            if pack_amt is not None:
                try:
                    amount = float(pack_amt) if isinstance(pack_amt, (int, float)) else None
                    income = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0.0
                    used_global_id = int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
                    
                    packages.append({
                        'package_number': pack_num,
                        'amount': amount,
                        'name': str(pack_name) if pack_name else 'N/A',
                        'income': income,
                        'used_global_id': used_global_id
                    })
                except:
                    pass
        
        if packages:
            key = email.lower().strip() if email else (member_id.upper().strip() if member_id else None)
            if key:
                excel_packages[key] = {
                    'email': str(email).lower().strip() if email else None,
                    'member_id': str(member_id).upper().strip() if member_id else None,
                    'packages': packages
                }
    
    print(f"✅ Found packages for {len(excel_packages)} users in Excel")
    return excel_packages

def get_db_users_with_packages():
    """Get new users and their packages from database"""
    print("🔍 Fetching users and packages from database...")
    
    query = f"""
    SELECT 
        u.id,
        u.email,
        u.display_id,
        u.name,
        p.id as purchase_id,
        p.package_id,
        p.amount,
        p.income,
        p.effective_global_ids,
        p.active_until,
        p.status,
        CASE 
            WHEN p.income >= p.amount * 2 THEN 'EXPIRED'
            WHEN p.active_until < CURRENT_DATE THEN 'EXPIRED'
            ELSE 'ACTIVE'
        END as package_status,
        pk.name as package_name
    FROM users u
    LEFT JOIN purchases p ON u.id = p.user_id
    LEFT JOIN packages pk ON p.package_id = pk.id
    WHERE u.id >= {NEW_USER_MIN_ID} AND u.id <= {NEW_USER_MAX_ID}
    ORDER BY u.id, p.id
    """
    
    result = run_sql_query(query)
    if not result:
        return {}
    
    db_users = {}
    
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 14:
            user_id = parts[0].strip()
            email = parts[1].strip().lower() if parts[1] else None
            display_id = parts[2].strip() if parts[2] else None
            name = parts[3].strip() if parts[3] else None
            purchase_id = parts[4].strip() if parts[4] else None
            package_id = parts[5].strip() if parts[5] else None
            amount = float(parts[6].strip()) if parts[6] and parts[6].strip() != 'NULL' else None
            income = float(parts[7].strip()) if parts[7] and parts[7].strip() != 'NULL' else 0.0
            effective_global_ids = int(parts[8].strip()) if parts[8] and parts[8].strip() != 'NULL' and parts[8].strip().isdigit() else 0
            active_until = parts[9].strip() if parts[9] and parts[9].strip() != 'NULL' else None
            status = parts[10].strip() if parts[10] else None
            package_status = parts[11].strip() if parts[11] else None
            package_name = parts[12].strip() if parts[12] else None
            
            key = email if email else (display_id.upper() if display_id else None)
            if not key:
                continue
            
            if key not in db_users:
                db_users[key] = {
                    'user_id': user_id,
                    'email': email,
                    'display_id': display_id,
                    'name': name,
                    'packages': []
                }
            
            if purchase_id and package_id:
                db_users[key]['packages'].append({
                    'purchase_id': purchase_id,
                    'package_id': package_id,
                    'package_name': package_name,
                    'amount': amount,
                    'income': income,
                    'effective_global_ids': effective_global_ids,
                    'active_until': active_until,
                    'status': status,
                    'package_status': package_status
                })
    
    print(f"✅ Found {len(db_users)} users in database")
    return db_users

def cross_check():
    """Main cross-check function"""
    print("=" * 100)
    print("🔍 CROSS-CHECK: Packages for Migrated New Users")
    print("=" * 100)
    print()
    
    # Get Excel data
    excel_packages = get_excel_packages()
    
    # Get DB data
    db_users = get_db_users_with_packages()
    
    # Match and compare
    print("\n" + "=" * 100)
    print("📊 COMPARISON RESULTS")
    print("=" * 100)
    print()
    
    matches = []
    mismatches = []
    missing_in_db = []
    missing_in_excel = []
    
    # Check Excel users
    for key, excel_data in excel_packages.items():
        db_data = db_users.get(key)
        
        if not db_data:
            missing_in_db.append({
                'key': key,
                'excel': excel_data
            })
            continue
        
        excel_pkg_list = excel_data['packages']
        db_pkg_list = db_data['packages']
        
        # Match packages by amount
        matched_packages = []
        unmatched_excel = []
        unmatched_db = []
        
        for excel_pkg in excel_pkg_list:
            matched = False
            for db_pkg in db_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    matched_packages.append({
                        'excel': excel_pkg,
                        'db': db_pkg
                    })
                    break
            if not matched:
                unmatched_excel.append(excel_pkg)
        
        for db_pkg in db_pkg_list:
            matched = False
            for excel_pkg in excel_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    break
            if not matched:
                unmatched_db.append(db_pkg)
        
        # Check for mismatches
        has_mismatch = False
        mismatch_details = []
        
        for match in matched_packages:
            excel_pkg = match['excel']
            db_pkg = match['db']
            
            issues = []
            
            # Check income
            if abs(excel_pkg['income'] - db_pkg['income']) > 0.01:
                issues.append(f"Income: Excel={excel_pkg['income']:.2f} vs DB={db_pkg['income']:.2f}")
            
            # Check global IDs
            if excel_pkg['used_global_id'] != db_pkg['effective_global_ids']:
                issues.append(f"Global IDs: Excel={excel_pkg['used_global_id']} vs DB={db_pkg['effective_global_ids']}")
            
            if issues:
                has_mismatch = True
                mismatch_details.append({
                    'package': excel_pkg,
                    'db_package': db_pkg,
                    'issues': issues
                })
        
        if unmatched_excel or unmatched_db or has_mismatch:
            mismatches.append({
                'user': db_data['display_id'] or db_data['email'],
                'name': db_data['name'],
                'matched': matched_packages,
                'unmatched_excel': unmatched_excel,
                'unmatched_db': unmatched_db,
                'mismatch_details': mismatch_details
            })
        else:
            matches.append({
                'user': db_data['display_id'] or db_data['email'],
                'name': db_data['name'],
                'packages': matched_packages
            })
    
    # Check DB users not in Excel
    for key, db_data in db_users.items():
        if key not in excel_packages:
            missing_in_excel.append({
                'key': key,
                'db': db_data
            })
    
    # Print results
    print(f"✅ Perfect Matches: {len(matches)}")
    print(f"⚠️  Mismatches/Issues: {len(mismatches)}")
    print(f"❌ Missing in DB: {len(missing_in_db)}")
    print(f"❌ Missing in Excel: {len(missing_in_excel)}")
    print()
    
    # Print perfect matches
    if matches:
        print("\n" + "=" * 100)
        print("✅ PERFECT MATCHES")
        print("=" * 100)
        for match in matches:
            print(f"\n👤 {match['user']} - {match['name']}")
            for pkg in match['packages']:
                excel = pkg['excel']
                db = pkg['db']
                print(f"  📦 {excel['name']} (₹{excel['amount']:,.2f})")
                print(f"     Income: ₹{excel['income']:.2f} | Global IDs: {excel['used_global_id']} | Status: {db['package_status']}")
    
    # Print mismatches
    if mismatches:
        print("\n" + "=" * 100)
        print("⚠️  MISMATCHES / ISSUES")
        print("=" * 100)
        for mismatch in mismatches:
            print(f"\n👤 {mismatch['user']} - {mismatch['name']}")
            
            # Matched packages with issues
            for detail in mismatch['mismatch_details']:
                excel = detail['package']
                db = detail['db_package']
                print(f"  📦 {excel['name']} (₹{excel['amount']:,.2f})")
                print(f"     Excel: Income=₹{excel['income']:.2f}, Global IDs={excel['used_global_id']}")
                print(f"     DB:    Income=₹{db['income']:.2f}, Global IDs={db['effective_global_ids']}, Status={db['package_status']}")
                for issue in detail['issues']:
                    print(f"     ⚠️  {issue}")
            
            # Unmatched Excel packages
            if mismatch['unmatched_excel']:
                print(f"  📦 Packages in Excel but NOT in DB:")
                for pkg in mismatch['unmatched_excel']:
                    print(f"     - {pkg['name']} (₹{pkg['amount']:,.2f})")
            
            # Unmatched DB packages
            if mismatch['unmatched_db']:
                print(f"  📦 Packages in DB but NOT in Excel:")
                for pkg in mismatch['unmatched_db']:
                    print(f"     - {pkg['package_name']} (₹{pkg['amount']:,.2f})")
    
    # Print missing in DB
    if missing_in_db:
        print("\n" + "=" * 100)
        print("❌ USERS IN EXCEL BUT NOT IN DB")
        print("=" * 100)
        for item in missing_in_db:
            print(f"\n👤 {item['key']}")
            for pkg in item['excel']['packages']:
                print(f"  📦 {pkg['name']} (₹{pkg['amount']:,.2f})")
    
    # Print missing in Excel
    if missing_in_excel:
        print("\n" + "=" * 100)
        print("❌ USERS IN DB BUT NOT IN EXCEL")
        print("=" * 100)
        for item in missing_in_excel:
            print(f"\n👤 {item['db']['display_id']} - {item['db']['name']}")
            for pkg in item['db']['packages']:
                print(f"  📦 {pkg['package_name']} (₹{pkg['amount']:,.2f})")
    
    # Summary statistics
    print("\n" + "=" * 100)
    print("📊 SUMMARY STATISTICS")
    print("=" * 100)
    
    total_excel_packages = sum(len(data['packages']) for data in excel_packages.values())
    total_db_packages = sum(len(data['packages']) for data in db_users.values())
    
    active_packages = 0
    expired_packages = 0
    
    for db_data in db_users.values():
        for pkg in db_data['packages']:
            if pkg['package_status'] == 'ACTIVE':
                active_packages += 1
            else:
                expired_packages += 1
    
    print(f"Total Packages in Excel: {total_excel_packages}")
    print(f"Total Packages in DB: {total_db_packages}")
    print(f"Active Packages in DB: {active_packages}")
    print(f"Expired Packages in DB: {expired_packages}")
    print()
    
    # Save detailed report
    report = {
        'verification_date': datetime.now().isoformat(),
        'summary': {
            'perfect_matches': len(matches),
            'mismatches': len(mismatches),
            'missing_in_db': len(missing_in_db),
            'missing_in_excel': len(missing_in_excel),
            'total_excel_packages': total_excel_packages,
            'total_db_packages': total_db_packages,
            'active_packages': active_packages,
            'expired_packages': expired_packages
        },
        'matches': matches,
        'mismatches': mismatches,
        'missing_in_db': missing_in_db,
        'missing_in_excel': missing_in_excel
    }
    
    with open('package-cross-check-report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    print(f"📄 Detailed report saved to: package-cross-check-report.json")

if __name__ == '__main__':
    cross_check()

