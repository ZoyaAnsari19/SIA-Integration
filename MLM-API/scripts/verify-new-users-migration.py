#!/usr/bin/env python3
"""
Verification script for new users migration
Compares local database with Excel data
"""

import openpyxl
import subprocess
import json
import re
from datetime import datetime

# Configuration
EXCEL_FILE = 'products-export-3.xlsx'
USE_LOCAL_DB = True
LOCAL_CONTAINER = 'mlm-postgres-local'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    # Clean query
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\"'\"'")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'psql', '-U', DB_USER, '-d', DB_NAME, '-t', '-A', '-F', '|', '-c', escaped_query
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    
    if result.returncode != 0:
        return None
    
    return result.stdout.strip()

def get_excel_users():
    """Read all users from Excel file"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return []
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    users = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        user_id = ws.cell(row_idx, headers.get('ID', 0)).value if headers.get('ID') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        
        if not email and not member_id:
            continue
        
        # Get package data
        packages = []
        for pack_num in range(1, 6):
            pack_name_col = f'Pack. {pack_num} name'
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_income_col = f'Pack. {pack_num} self + global'
            pack_global_col = f'Pack. {pack_num} used global id'
            
            pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else None
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else None
            pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else None
            
            if pack_name or pack_amt:
                packages.append({
                    'name': str(pack_name) if pack_name else None,
                    'amount': float(pack_amt) if pack_amt and isinstance(pack_amt, (int, float)) else None,
                    'income': float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0.0,
                    'used_global_ids': int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
                })
        
        # Get wallet data
        spot_amt_col = 'Spot amt.'
        other_amt_col = 'Other amt.'
        spot_balance = ws.cell(row_idx, headers.get(spot_amt_col, 0)).value if headers.get(spot_amt_col) else 0
        other_balance = ws.cell(row_idx, headers.get(other_amt_col, 0)).value if headers.get(other_amt_col) else 0
        
        users.append({
            'excel_id': str(user_id) if user_id else None,
            'member_id': str(member_id) if member_id else None,
            'email': str(email).lower().strip() if email else None,
            'name': ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None,
            'phone': ws.cell(row_idx, headers.get('Phone', 0)).value if headers.get('Phone') else None,
            'sponsor_id': ws.cell(row_idx, headers.get('Sponsor ID', 0)).value if headers.get('Sponsor ID') else None,
            'packages': packages,
            'spot_balance': float(spot_balance) if spot_balance and isinstance(spot_balance, (int, float)) else 0.0,
            'other_balance': float(other_balance) if other_balance and isinstance(other_balance, (int, float)) else 0.0
        })
    
    print(f"✅ Found {len(users)} users in Excel")
    return users

def get_db_users():
    """Get all users from database"""
    print("🔍 Fetching users from database...")
    query = """
    SELECT 
        id, email, display_id, name, phone, referrer_user_id, status, created_at
    FROM users
    ORDER BY id DESC
    LIMIT 100
    """
    
    result = run_sql_query(query)
    if not result:
        return []
    
    users = []
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 8:
            users.append({
                'id': parts[0].strip(),
                'email': parts[1].strip().lower() if parts[1] else None,
                'display_id': parts[2].strip() if parts[2] else None,
                'name': parts[3].strip() if parts[3] else None,
                'phone': parts[4].strip() if parts[4] else None,
                'referrer_user_id': parts[5].strip() if parts[5] else None,
                'status': parts[6].strip() if parts[6] else None,
                'created_at': parts[7].strip() if parts[7] else None
            })
    
    print(f"✅ Found {len(users)} recent users in DB")
    return users

def get_user_wallet(user_id):
    """Get user wallet balance"""
    query = f"""
    SELECT spot_balance, other_balance, balance
    FROM user_balances
    WHERE user_id = {user_id}
    LIMIT 1
    """
    
    result = run_sql_query(query)
    if not result:
        return None
    
    parts = result.split('|')
    if len(parts) >= 3:
        return {
            'spot_balance': float(parts[0].strip()) if parts[0] else 0.0,
            'other_balance': float(parts[1].strip()) if parts[1] else 0.0,
            'balance': float(parts[2].strip()) if parts[2] else 0.0
        }
    return None

def get_user_purchases(user_id):
    """Get user purchases"""
    query = f"""
    SELECT 
        id, package_id, amount, income, effective_global_ids, 
        active_until, status, created_at
    FROM purchases
    WHERE user_id = {user_id}
    ORDER BY created_at DESC
    """
    
    result = run_sql_query(query)
    if not result:
        return []
    
    purchases = []
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 8:
            purchases.append({
                'id': parts[0].strip(),
                'package_id': parts[1].strip() if parts[1] else None,
                'amount': float(parts[2].strip()) if parts[2] else 0.0,
                'income': float(parts[3].strip()) if parts[3] else 0.0,
                'effective_global_ids': int(parts[4].strip()) if parts[4] and parts[4].strip().isdigit() else 0,
                'active_until': parts[5].strip() if parts[5] else None,
                'status': parts[6].strip() if parts[6] else None,
                'created_at': parts[7].strip() if parts[7] else None
            })
    
    return purchases

def get_user_tree_paths(user_id):
    """Get user tree paths count"""
    query = f"""
    SELECT COUNT(*) 
    FROM user_tree_paths
    WHERE descendant_id = {user_id}
    """
    
    result = run_sql_query(query)
    if result and result.strip().isdigit():
        return int(result.strip())
    return 0

def verify_migration():
    """Main verification function"""
    print("=" * 80)
    print("🔍 VERIFICATION: New Users Migration")
    print("=" * 80)
    print()
    
    # Get Excel users
    excel_users = get_excel_users()
    
    # Get recent DB users (last 100)
    db_users = get_db_users()
    
    # Find recently created users (check by highest IDs - likely migrated users)
    # Get highest user IDs (assuming new users have highest IDs)
    max_id_query = "SELECT MAX(id) FROM users;"
    max_id_result = run_sql_query(max_id_query)
    max_id = int(max_id_result.strip()) if max_id_result and max_id_result.strip().isdigit() else 0
    
    print(f"\n📊 Highest user ID in DB: {max_id}")
    print(f"📊 Total users in DB: {len(db_users)}")
    
    # Check users with IDs close to max (likely new users)
    # Assuming new users are in the last 100 IDs
    recent_users = []
    threshold_id = max_id - 100 if max_id > 100 else 0
    
    for db_user in db_users:
        user_id = int(db_user['id']) if db_user['id'].isdigit() else 0
        if user_id >= threshold_id:
            recent_users.append(db_user)
    
    print(f"🆕 Users with ID >= {threshold_id}: {len(recent_users)}")
    
    # Match Excel users with DB users
    # Focus on users with high IDs (1900+) which are likely migrated
    verification_results = []
    
    # Get all DB users with high IDs
    high_id_query = f"""
    SELECT id, email, display_id, name, phone, referrer_user_id, status, created_at
    FROM users
    WHERE id >= {threshold_id}
    ORDER BY id DESC
    """
    high_id_result = run_sql_query(high_id_query)
    high_id_users = []
    if high_id_result:
        for line in high_id_result.split('\n'):
            if not line.strip():
                continue
            parts = line.split('|')
            if len(parts) >= 8:
                high_id_users.append({
                    'id': parts[0].strip(),
                    'email': parts[1].strip().lower() if parts[1] else None,
                    'display_id': parts[2].strip() if parts[2] else None,
                    'name': parts[3].strip() if parts[3] else None,
                    'phone': parts[4].strip() if parts[4] else None,
                    'referrer_user_id': parts[5].strip() if parts[5] else None,
                    'status': parts[6].strip() if parts[6] else None,
                    'created_at': parts[7].strip() if parts[7] else None
                })
    
    print(f"🔍 Found {len(high_id_users)} users with ID >= {threshold_id}")
    
    # Match Excel users with high ID DB users
    for db_user in high_id_users:
        # Find matching Excel user
        excel_user = None
        for eu in excel_users:
            if db_user['email'] and eu['email'] and db_user['email'] == eu['email']:
                excel_user = eu
                break
            elif db_user['display_id'] and eu['member_id'] and db_user['display_id'].upper() == eu['member_id'].upper():
                excel_user = eu
                break
        
        if not excel_user:
            continue
        
        # Get wallet
        wallet = get_user_wallet(db_user['id'])
        
        # Get purchases
        purchases = get_user_purchases(db_user['id'])
        
        # Get tree paths
        tree_paths_count = get_user_tree_paths(db_user['id'])
        
        # Compare with Excel
        excel_packages = [p for p in excel_user['packages'] if p.get('amount')]
        
        verification_results.append({
            'user_id': db_user['id'],
            'display_id': db_user['display_id'],
            'email': db_user['email'],
            'name': db_user['name'],
            'excel_packages_count': len(excel_packages),
            'db_purchases_count': len(purchases),
            'excel_spot': excel_user['spot_balance'],
            'db_spot': wallet['spot_balance'] if wallet else 0.0,
            'excel_other': excel_user['other_balance'],
            'db_other': wallet['other_balance'] if wallet else 0.0,
            'tree_paths': tree_paths_count,
            'packages_match': len(excel_packages) == len(purchases),
            'wallet_match': wallet and abs(excel_user['spot_balance'] - wallet['spot_balance']) < 0.01 and abs(excel_user['other_balance'] - wallet['other_balance']) < 0.01,
            'purchases': purchases,
            'excel_packages': excel_packages
        })
    
    # Print results
    print("\n" + "=" * 80)
    print("📊 VERIFICATION RESULTS")
    print("=" * 80)
    print()
    
    if not verification_results:
        print("⚠️  No recently migrated users found to verify")
        return
    
    print(f"✅ Verified {len(verification_results)} migrated users\n")
    
    issues = []
    success_count = 0
    
    for result in verification_results:
        status = "✅"
        issues_list = []
        
        if not result['wallet_match']:
            status = "❌"
            issues_list.append(f"Wallet mismatch: Excel(Spot={result['excel_spot']:.2f}, Other={result['excel_other']:.2f}) vs DB(Spot={result['db_spot']:.2f}, Other={result['db_other']:.2f})")
        
        if not result['packages_match']:
            status = "⚠️"
            issues_list.append(f"Package count mismatch: Excel={result['excel_packages_count']} vs DB={result['db_purchases_count']}")
        
        if result['tree_paths'] == 0:
            status = "❌"
            issues_list.append("No tree paths created")
        
        if status == "✅":
            success_count += 1
        
        if issues_list:
            issues.append({
                'user': f"{result['display_id']} ({result['email']})",
                'issues': issues_list
            })
        
        print(f"{status} {result['display_id']} - {result['name']}")
        print(f"   Packages: Excel={result['excel_packages_count']}, DB={result['db_purchases_count']}")
        print(f"   Wallet: Excel(Spot={result['excel_spot']:.2f}, Other={result['excel_other']:.2f}) | DB(Spot={result['db_spot']:.2f}, Other={result['db_other']:.2f})")
        print(f"   Tree Paths: {result['tree_paths']}")
        
        # Show purchase details
        if result['purchases']:
            for purchase in result['purchases']:
                print(f"   📦 Purchase: Amount={purchase['amount']:.2f}, Income={purchase['income']:.2f}, Global IDs={purchase['effective_global_ids']}")
        print()
    
    print("=" * 80)
    print(f"✅ Successfully verified: {success_count}/{len(verification_results)}")
    if issues:
        print(f"⚠️  Issues found: {len(issues)}")
        print("\nIssues:")
        for issue in issues:
            print(f"  - {issue['user']}:")
            for i in issue['issues']:
                print(f"    • {i}")
    
    # Save detailed report
    report = {
        'verification_date': datetime.now().isoformat(),
        'total_verified': len(verification_results),
        'successful': success_count,
        'issues_count': len(issues),
        'results': verification_results,
        'issues': issues
    }
    
    with open('new-users-verification-report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    print(f"\n📄 Detailed report saved to: new-users-verification-report.json")

if __name__ == '__main__':
    verify_migration()

