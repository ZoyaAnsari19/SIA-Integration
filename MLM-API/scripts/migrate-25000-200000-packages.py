#!/usr/bin/env python3
"""
Migration script for ₹25,000 and ₹200,000 package users from products-export-3.xlsx

Process:
1. Update wallet balances (vanish + set Excel amounts)
2. Map package (₹25,000 → ID 12, ₹200,000 → ID 13)
3. Create purchase record with Excel income
4. Check package expiry (2x amount)
5. Check if reinvestment
6. Process SPOT commissions (skip direct referrer, skip qualified uplines, add unqualified to pending)
"""

import openpyxl
import subprocess
import json
from datetime import datetime, timedelta
from decimal import Decimal

# Configuration
EXCEL_FILE = 'products-export-3.xlsx'
KUBECONFIG = 'azure-kube/0ffdcdf4-849b-4521-9868-be1000865e08'
NAMESPACE = 'mlm'
POD_NAME = 'postgres-0'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

# Package mapping
PACKAGE_MAPPING = {
    25000: 12,   # ₹25,000 → Package ID 12
    200000: 13   # ₹200,000 → Package ID 13
}

TARGET_AMOUNTS = [25000, 200000]

def run_sql_query(query, description=""):
    """Execute SQL query via kubectl exec"""
    if description:
        print(f"    {description}...")
    
    # Clean query (remove extra whitespace, newlines)
    query = ' '.join(query.split())
    
    # Escape single quotes in query for shell
    escaped_query = query.replace("'", "'\"'\"'")
    
    cmd = [
        'kubectl', '--kubeconfig', KUBECONFIG,
        'exec', '-n', NAMESPACE, POD_NAME, '--',
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=60
    )
    
    if result.returncode != 0:
        if description:
            error_msg = result.stderr[:200] if result.stderr else "Unknown error"
            print(f"    ❌ Error: {error_msg}")
        return None
    
    return result.stdout.strip()

def get_user_from_db(email=None, display_id=None, user_id=None):
    """Find user in database by email, display_id, or user_id"""
    conditions = []
    params = []
    
    if email:
        conditions.append("LOWER(email) = LOWER(%s)")
        params.append(email)
    if display_id:
        conditions.append("LOWER(display_id) = LOWER(%s)")
        params.append(display_id)
    if user_id:
        conditions.append("id = %s")
        params.append(user_id)
    
    if not conditions:
        return None
    
    query = f"""
    SELECT id, email, display_id, name, referrer_user_id
    FROM users
    WHERE {' OR '.join(conditions)}
    LIMIT 1;
    """
    
    # For kubectl exec, we need to format the query differently
    # Let's use a simpler approach with direct SQL
    if email:
        escaped_email = email.replace("'", "''")
        query = f"SELECT id, email, display_id, name, referrer_user_id FROM users WHERE LOWER(email) = LOWER('{escaped_email}') LIMIT 1;"
        result = run_sql_query(query, f"Finding user by email: {email}")
        if result:
            parts = result.split('|')
            if len(parts) >= 5:
                return {
                    'id': parts[0].strip(),
                    'email': parts[1].strip(),
                    'display_id': parts[2].strip() if parts[2] else None,
                    'name': parts[3].strip() if parts[3] else None,
                    'referrer_user_id': parts[4].strip() if parts[4] else None
                }
    
    if display_id:
        escaped_display_id = display_id.replace("'", "''")
        query = f"SELECT id, email, display_id, name, referrer_user_id FROM users WHERE LOWER(display_id) = LOWER('{escaped_display_id}') LIMIT 1;"
        result = run_sql_query(query, f"Finding user by display_id: {display_id}")
        if result:
            parts = result.split('|')
            if len(parts) >= 5:
                return {
                    'id': parts[0].strip(),
                    'email': parts[1].strip(),
                    'display_id': parts[2].strip() if parts[2] else None,
                    'name': parts[3].strip() if parts[3] else None,
                    'referrer_user_id': parts[4].strip() if parts[4] else None
                }
    
    if user_id:
        result = run_sql_query(
            f"SELECT id, email, display_id, name, referrer_user_id FROM users WHERE id = {user_id} LIMIT 1;",
            f"Finding user by id: {user_id}"
        )
        if result:
            parts = result.split('|')
            if len(parts) >= 5:
                return {
                    'id': parts[0].strip(),
                    'email': parts[1].strip(),
                    'display_id': parts[2].strip() if parts[2] else None,
                    'name': parts[3].strip() if parts[3] else None,
                    'referrer_user_id': parts[4].strip() if parts[4] else None
                }
    
    return None

def get_package_details(package_id):
    """Get package details from database"""
    query = f"""
    SELECT id, name, price, global_ids, validity_months
    FROM packages
    WHERE id = {package_id};
    """
    
    result = run_sql_query(query, f"Getting package {package_id} details")
    if result:
        parts = result.split('|')
        if len(parts) >= 5:
            return {
                'id': parts[0].strip(),
                'name': parts[1].strip() if parts[1] else None,
                'price': float(parts[2]) if parts[2] else 0,
                'global_ids': int(parts[3]) if parts[3] else None,
                'validity_months': int(parts[4]) if parts[4] else 12
            }
    return None

def check_if_reinvestment(user_id, purchase_date):
    """Check if this is a reinvestment (user has active packages before purchase date)"""
    query = f"SELECT COUNT(*) FROM purchases WHERE user_id = {user_id} AND status = 'completed' AND purchased_at < '{purchase_date}' AND active_until >= CURRENT_DATE;"
    
    result = run_sql_query(query)
    if result:
        try:
            count = int(result.strip())
            if count > 0:
                # Check if any previous purchase has not reached 2x
                query_2x = f"SELECT id, amount, income FROM purchases WHERE user_id = {user_id} AND status = 'completed' AND purchased_at < '{purchase_date}' AND active_until >= CURRENT_DATE;"
                
                result_2x = run_sql_query(query_2x)
                if result_2x:
                    for line in result_2x.split('\n'):
                        if line.strip():
                            parts = line.split('|')
                            if len(parts) >= 3:
                                try:
                                    amount = float(parts[1]) if parts[1] else 0
                                    income = float(parts[2]) if parts[2] else 0
                                    
                                    # Check if not reached 2x
                                    if income < amount * 2:
                                        return True
                                except:
                                    pass
        except:
            pass
    
    return False

def check_eligibility(user_id, level):
    """Check if user is eligible for a specific level"""
    query = f"SELECT eligibility::text FROM level_eligibility WHERE user_id = {user_id};"
    
    result = run_sql_query(query, f"Checking eligibility for user {user_id}, level {level}")
    if result:
        try:
            # Parse JSON eligibility
            eligibility = json.loads(result)
            return bool(eligibility.get(str(level), False))
        except Exception as e:
            # Try alternative parsing
            try:
                # Sometimes the result might have extra formatting
                cleaned = result.strip().strip('"').replace("\\\"", '"')
                eligibility = json.loads(cleaned)
                return bool(eligibility.get(str(level), False))
            except:
                pass
    
    return False

def get_spot_percent(level):
    """Get SPOT commission percentage for a level"""
    query = f"""
    SELECT spot_commission_percent
    FROM levels
    WHERE level = {level};
    """
    
    result = run_sql_query(query, f"Getting SPOT percent for level {level}")
    if result:
        try:
            percent = float(result.strip())
            return percent
        except:
            pass
    
    # Fallback to commission_rules
    query_fallback = f"""
    SELECT percent
    FROM commission_rules
    WHERE type = 'LEVEL_SPOT' AND level = {level}
    LIMIT 1;
    """
    
    result = run_sql_query(query_fallback)
    if result:
        try:
            percent = float(result.strip())
            return percent
        except:
            pass
    
    return 0.0

def get_uplines(user_id, max_depth=10):
    """Get upline chain for a user"""
    query = f"""
    SELECT ancestor_id, depth
    FROM user_tree_paths
    WHERE descendant_id = {user_id}
      AND depth >= 1
      AND depth <= {max_depth}
    ORDER BY depth ASC;
    """
    
    result = run_sql_query(query, f"Getting uplines for user {user_id}")
    uplines = []
    
    if result:
        for line in result.split('\n'):
            if line.strip():
                parts = line.split('|')
                if len(parts) >= 2:
                    uplines.append({
                        'ancestor_id': parts[0].strip(),
                        'depth': int(parts[1].strip())
                    })
    
    return uplines

def main():
    print("=" * 120)
    print("🚀 MIGRATION SCRIPT - ₹25,000 and ₹200,000 Package Users")
    print("=" * 120)
    print()
    
    # Step 1: Read Excel file
    print("📄 Step 1: Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    # Extract users with target packages
    excel_users = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data[header] = cell_value
        
        user_id = str(row_data.get('ID', '')) if row_data.get('ID') else None
        member_id = str(row_data.get('Member ID', '')) if row_data.get('Member ID') else None
        user_name = str(row_data.get('Name', '')) if row_data.get('Name') else None
        user_email = str(row_data.get('Email', '')) if row_data.get('Email') else None
        
        # Get wallet amounts from Excel
        spot_amt = row_data.get('Spot amt.')
        other_amt = row_data.get('Other amt.')
        
        excel_spot = float(spot_amt) if spot_amt is not None else 0
        excel_other = float(other_amt) if other_amt is not None else 0
        
        # Check packages
        for i in range(1, 9):
            pkg_amt_col = f'Pack. {i} amt.'
            pkg_name_col = f'Pack. {i} name'
            pkg_income_col = f'Pack. {i} self + global'
            pkg_used_global_id_col = f'Pack. {i} used global id'
            
            pkg_amount = row_data.get(pkg_amt_col)
            pkg_name = row_data.get(pkg_name_col)
            pkg_income = row_data.get(pkg_income_col)
            pkg_used_global_id = row_data.get(pkg_used_global_id_col)
            
            if pkg_amount is not None:
                try:
                    amount_float = float(pkg_amount)
                    if any(abs(amount_float - target) < 1.0 for target in TARGET_AMOUNTS):
                        income_float = float(pkg_income) if pkg_income is not None else 0
                        used_global_id = int(pkg_used_global_id) if pkg_used_global_id is not None and str(pkg_used_global_id) != 'N/A' else None
                        
                        excel_users.append({
                            'user_id': user_id,
                            'member_id': member_id,
                            'name': user_name,
                            'email': user_email,
                            'package_amount': amount_float,
                            'package_name': str(pkg_name) if pkg_name else 'N/A',
                            'excel_income': income_float,
                            'excel_spot': excel_spot,
                            'excel_other': excel_other,
                            'excel_used_global_id': used_global_id,
                            'package_number': i
                        })
                        break
                except:
                    pass
    
    print(f"✅ Found {len(excel_users)} users with ₹25,000 or ₹200,000 packages")
    print()
    
    # Migration summary
    migration_summary = {
        'total_users': len(excel_users),
        'wallets_updated': 0,
        'purchases_created': 0,
        'expired_packages': 0,
        'active_packages': 0,
        'pending_commissions_created': 0,
        'errors': []
    }
    
    # Step 2: Process each user
    print("=" * 120)
    print("📦 Step 2: Processing Users...")
    print("=" * 120)
    print()
    
    for idx, excel_user in enumerate(excel_users, 1):
        print(f"\n[{idx}/{len(excel_users)}] Processing: {excel_user['name']} ({excel_user['email']})")
        print(f"  Package: ₹{excel_user['package_amount']:,.2f} | Income: ₹{excel_user['excel_income']:,.2f}")
        
        # Find user in database
        db_user = get_user_from_db(
            email=excel_user['email'],
            display_id=excel_user['member_id'],
            user_id=excel_user['user_id']
        )
        
        if not db_user:
            error_msg = f"User not found in database: {excel_user['email']}"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': excel_user['name'],
                'error': error_msg
            })
            continue
        
        print(f"  ✅ Found in DB: ID {db_user['id']}, Name: {db_user['name']}")
        
        # Step 3: Update wallet balances
        print(f"  💰 Updating wallet balances...")
        wallet_query = f"""
        UPDATE user_balances
        SET 
          spot_balance = {excel_user['excel_spot']},
          other_balance = {excel_user['excel_other']},
          balance = {excel_user['excel_spot'] + excel_user['excel_other']},
          updated_at = NOW()
        WHERE user_id = {db_user['id']};
        """
        
        result = run_sql_query(wallet_query, "Updating wallet")
        if result is not None:
            print(f"  ✅ Wallet updated: Spot ₹{excel_user['excel_spot']:,.2f}, Other ₹{excel_user['excel_other']:,.2f}")
            migration_summary['wallets_updated'] += 1
        else:
            print(f"  ❌ Failed to update wallet")
            continue
        
        # Step 4: Map package
        package_id = PACKAGE_MAPPING.get(int(excel_user['package_amount']))
        if not package_id:
            error_msg = f"Package mapping not found for amount: {excel_user['package_amount']}"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': excel_user['name'],
                'error': error_msg
            })
            continue
        
        package = get_package_details(package_id)
        if not package:
            error_msg = f"Package {package_id} not found in database"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': excel_user['name'],
                'error': error_msg
            })
            continue
        
        print(f"  ✅ Package mapped: {package['name']} (ID: {package_id})")
        
        # Step 5: Create purchase record
        print(f"  📦 Creating purchase record...")
        
        # Calculate active_until (initially future date - display only)
        purchased_at = datetime.now()
        active_until = purchased_at + timedelta(days=365)  # 1 year from now (display only)
        
        # Check if expired (2x amount)
        if excel_user['excel_income'] >= excel_user['package_amount'] * 2:
            active_until = purchased_at  # Mark as expired
            print(f"  ⚠️  Package EXPIRED (income >= 2x amount)")
            migration_summary['expired_packages'] += 1
        else:
            print(f"  ✅ Package ACTIVE (income < 2x amount)")
            migration_summary['active_packages'] += 1
        
        # Format dates for PostgreSQL
        purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
        active_until_str = active_until.strftime('%Y-%m-%d %H:%M:%S')
        
        # Create purchase
        # Use Excel's used global id if available, otherwise use package.global_ids
        if excel_user.get('excel_used_global_id') is not None:
            global_ids_value = excel_user['excel_used_global_id']
            print(f"  📊 Using Excel used global ID: {global_ids_value}")
        else:
            global_ids_value = package['global_ids'] if package['global_ids'] else 'NULL'
            print(f"  📊 Using package global IDs: {global_ids_value}")
        
        global_ids_sql = global_ids_value if global_ids_value != 'NULL' else 'NULL'
        purchase_query = f"""
        INSERT INTO purchases (
          user_id,
          package_id,
          purchase_type,
          amount,
          purchased_at,
          active_until,
          status,
          is_manual,
          is_renewal,
          income,
          effective_global_ids
        ) VALUES (
          {db_user['id']},
          {package_id},
          'DIRECT_PACKAGE',
          {excel_user['package_amount']},
          '{purchased_at_str}',
          '{active_until_str}',
          'completed',
          true,
          false,
          {excel_user['excel_income']},
          {global_ids_sql}
        )
        RETURNING id;
        """
        
        result = run_sql_query(purchase_query, "Creating purchase")
        if not result:
            error_msg = "Failed to create purchase"
            print(f"  ❌ {error_msg}")
            migration_summary['errors'].append({
                'user': excel_user['name'],
                'error': error_msg
            })
            continue
        
        purchase_id = result.strip()
        print(f"  ✅ Purchase created: ID {purchase_id}")
        migration_summary['purchases_created'] += 1
        
        # Step 6: Check if reinvestment
        purchased_at_str = purchased_at.strftime('%Y-%m-%d %H:%M:%S')
        is_reinvestment = check_if_reinvestment(db_user['id'], purchased_at_str)
        if is_reinvestment:
            print(f"  🔄 Reinvestment detected (50% reduction for Level 1+)")
        else:
            print(f"  ✨ First purchase (100% SPOT)")
        
        # Step 7: Process SPOT commissions
        print(f"  💵 Processing SPOT commissions...")
        
        uplines = get_uplines(db_user['id'], max_depth=10)
        print(f"  📊 Found {len(uplines)} uplines")
        
        pending_count = 0
        skipped_qualified = 0
        skipped_direct = 0
        
        for upline in uplines:
            depth = upline['depth']
            level = depth - 1  # depth 2 → level 1, etc.
            upline_id = upline['ancestor_id']
            
            # Skip direct referrer (Level 0, depth 1)
            if depth == 1:
                skipped_direct += 1
                continue
            
            # For Team Uplines (Level 1-9, depth 2-10)
            if depth >= 2 and depth <= 10:
                # Get SPOT percentage
                spot_percent = get_spot_percent(level)
                if spot_percent == 0:
                    continue
                
                # Calculate SPOT amount
                spot_amount = excel_user['package_amount'] * spot_percent / 100
                
                # Apply 50% reduction for reinvestments (Level 1+)
                if is_reinvestment and level >= 1:
                    spot_amount = spot_amount * 0.5
                
                # Check eligibility
                eligible = check_eligibility(upline_id, level)
                
                if eligible:
                    # Skip qualified upline
                    skipped_qualified += 1
                    continue
                else:
                    # Add to pending_commissions
                    pending_query = f"""
                    INSERT INTO pending_commissions (
                      receiver_user_id,
                      source_user_id,
                      purchase_id,
                      level,
                      commission_type,
                      amount,
                      metadata
                    ) VALUES (
                      {upline_id},
                      {db_user['id']},
                      {purchase_id},
                      {level},
                      'SPOT',
                      {spot_amount},
                      '{{"level": {level}, "depth": {depth}, "reason": "eligibility", "is_reinvestment": {str(is_reinvestment).lower()}}}'::jsonb
                    )
                    ON CONFLICT DO NOTHING;
                    """
                    
                    result = run_sql_query(pending_query)
                    if result is not None:
                        pending_count += 1
                        migration_summary['pending_commissions_created'] += 1
        
        print(f"  ✅ SPOT processing complete:")
        print(f"     - Direct referrers skipped: {skipped_direct}")
        print(f"     - Qualified uplines skipped: {skipped_qualified}")
        print(f"     - Pending commissions created: {pending_count}")
    
    # Final summary
    print()
    print("=" * 120)
    print("📊 MIGRATION SUMMARY")
    print("=" * 120)
    print(f"Total Users Processed: {migration_summary['total_users']}")
    print(f"✅ Wallets Updated: {migration_summary['wallets_updated']}")
    print(f"✅ Purchases Created: {migration_summary['purchases_created']}")
    print(f"✅ Active Packages: {migration_summary['active_packages']}")
    print(f"⚠️  Expired Packages: {migration_summary['expired_packages']}")
    print(f"✅ Pending Commissions Created: {migration_summary['pending_commissions_created']}")
    print(f"❌ Errors: {len(migration_summary['errors'])}")
    
    if migration_summary['errors']:
        print("\n❌ ERRORS:")
        for error in migration_summary['errors']:
            print(f"  - {error['user']}: {error['error']}")
    
    # Save summary to file
    with open('migration-summary.json', 'w') as f:
        json.dump(migration_summary, f, indent=2, default=str)
    
    print("\n✅ Migration summary saved to migration-summary.json")
    print("=" * 120)

if __name__ == '__main__':
    main()

