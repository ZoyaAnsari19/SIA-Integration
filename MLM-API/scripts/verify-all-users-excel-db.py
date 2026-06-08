#!/usr/bin/env python3
"""
Comprehensive verification: All users in Excel vs Database
Checks: Users, Wallets, Packages, Income, Global IDs
"""

import openpyxl
import subprocess
import json
from datetime import datetime

# Configuration
EXCEL_FILE = 'products-export-3.xlsx'
USE_LOCAL_DB = True
LOCAL_CONTAINER = 'mlm-postgres-local'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        return None
    
    return result.stdout.strip()

def get_all_excel_users():
    """Read all users from Excel"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return {}
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    excel_users = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        user_id = ws.cell(row_idx, headers.get('ID', 0)).value if headers.get('ID') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
        
        if not email and not member_id:
            continue
        
        # Get wallet amounts
        spot_amt = ws.cell(row_idx, headers.get('Spot amt.', 0)).value if headers.get('Spot amt.') else 0
        other_amt = ws.cell(row_idx, headers.get('Other amt.', 0)).value if headers.get('Other amt.') else 0
        spot_balance = float(spot_amt) if spot_amt and isinstance(spot_amt, (int, float)) else 0.0
        other_balance = float(other_amt) if other_amt and isinstance(other_amt, (int, float)) else 0.0
        
        # Get packages
        packages = []
        for pack_num in range(1, 9):
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_name_col = f'Pack. {pack_num} name'
            pack_income_col = f'Pack. {pack_num} self + global'
            pack_global_col = f'Pack. {pack_num} used global id'
            
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            if pack_amt is not None:
                try:
                    amount = float(pack_amt) if isinstance(pack_amt, (int, float)) else None
                    pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
                    pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else 0
                    pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else 0
                    
                    income = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0.0
                    used_global_id = int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
                    
                    packages.append({
                        'amount': amount,
                        'name': str(pack_name) if pack_name else 'N/A',
                        'income': income,
                        'used_global_id': used_global_id
                    })
                except:
                    pass
        
        # Use display_id (Member ID) as primary key, email as secondary
        key = str(member_id).upper().strip() if member_id else (email.lower().strip() if email else None)
        if key:
            excel_users[key] = {
                'excel_id': str(user_id) if user_id else None,
                'member_id': str(member_id).upper().strip() if member_id else None,
                'email': str(email).lower().strip() if email else None,
                'name': str(name) if name else None,
                'spot_balance': spot_balance,
                'other_balance': other_balance,
                'packages': packages
            }
    
    print(f"✅ Found {len(excel_users)} users in Excel")
    return excel_users

def get_all_db_users():
    """Get all users from database with wallets and packages"""
    print("🔍 Fetching all users from database...")
    
    query = """
    SELECT 
        u.id,
        u.email,
        u.display_id,
        u.name,
        COALESCE(ub.spot_balance, 0) as spot_balance,
        COALESCE(ub.other_balance, 0) as other_balance,
        COALESCE(ub.balance, 0) as total_balance
    FROM users u
    LEFT JOIN user_balances ub ON u.id = ub.user_id
    ORDER BY u.id
    """
    
    result = run_sql_query(query)
    if not result:
        return {}
    
    db_users = {}
    
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 7:
            user_id = parts[0].strip()
            email = parts[1].strip().lower() if parts[1] else None
            display_id = parts[2].strip() if parts[2] else None
            name = parts[3].strip() if parts[3] else None
            spot_balance = float(parts[4].strip()) if parts[4] and parts[4].strip() != 'NULL' else 0.0
            other_balance = float(parts[5].strip()) if parts[5] and parts[5].strip() != 'NULL' else 0.0
            total_balance = float(parts[6].strip()) if parts[6] and parts[6].strip() != 'NULL' else 0.0
            
            # Use display_id as primary key, email as secondary
            key = display_id.upper() if display_id else (email if email else None)
            if not key:
                continue
            
            db_users[key] = {
                'user_id': user_id,
                'email': email,
                'display_id': display_id,
                'name': name,
                'spot_balance': spot_balance,
                'other_balance': other_balance,
                'total_balance': total_balance,
                'packages': []
            }
    
    # Get packages for all users
    print("🔍 Fetching packages from database...")
    query = """
    SELECT 
        u.email,
        u.display_id,
        p.amount,
        p.income,
        p.effective_global_ids,
        CASE 
            WHEN p.income >= p.amount * 2 THEN 'EXPIRED'
            ELSE 'ACTIVE'
        END as status,
        pk.name as package_name
    FROM purchases p
    JOIN users u ON p.user_id = u.id
    LEFT JOIN packages pk ON p.package_id = pk.id
    WHERE p.status = 'completed'
    ORDER BY u.id, p.id
    """
    
    result = run_sql_query(query)
    if result:
        for line in result.split('\n'):
            if not line.strip():
                continue
            parts = line.split('|')
            if len(parts) >= 7:
                email = parts[0].strip().lower() if parts[0] else None
                display_id = parts[1].strip() if parts[1] else None
                amount = float(parts[2].strip()) if parts[2] and parts[2].strip() != 'NULL' else 0.0
                income = float(parts[3].strip()) if parts[3] and parts[3].strip() != 'NULL' else 0.0
                effective_global_ids = int(parts[4].strip()) if parts[4] and parts[4].strip() != 'NULL' and parts[4].strip().isdigit() else 0
                status = parts[5].strip() if parts[5] else None
                package_name = parts[6].strip() if parts[6] else None
                
                # Use display_id as primary key, email as secondary
                key = display_id.upper() if display_id else (email if email else None)
                if key and key in db_users:
                    db_users[key]['packages'].append({
                        'amount': amount,
                        'income': income,
                        'effective_global_ids': effective_global_ids,
                        'status': status,
                        'name': package_name
                    })
    
    print(f"✅ Found {len(db_users)} users in database")
    return db_users

def verify_all():
    """Main verification function"""
    print("=" * 120)
    print("🔍 COMPREHENSIVE VERIFICATION: All Users (Excel vs Database)")
    print("=" * 120)
    print()
    
    # Get Excel data
    excel_users = get_all_excel_users()
    
    # Get DB data
    db_users = get_all_db_users()
    
    print("\n" + "=" * 120)
    print("📊 COMPARISON RESULTS")
    print("=" * 120)
    print()
    
    # Statistics
    excel_count = len(excel_users)
    db_count = len(db_users)
    
    print(f"📄 Total users in Excel: {excel_count}")
    print(f"💾 Total users in Database: {db_count}")
    print()
    
    # Match users
    matched_users = []
    missing_in_db = []
    missing_in_excel = []
    wallet_mismatches = []
    package_mismatches = []
    
    # Check Excel users
    for key, excel_data in excel_users.items():
        db_data = db_users.get(key)
        
        if not db_data:
            missing_in_db.append(excel_data)
            continue
        
        # Check wallet
        wallet_match = True
        wallet_issues = []
        
        if abs(excel_data['spot_balance'] - db_data['spot_balance']) > 0.01:
            wallet_match = False
            wallet_issues.append(f"Spot: Excel=₹{excel_data['spot_balance']:.2f} vs DB=₹{db_data['spot_balance']:.2f}")
        
        if abs(excel_data['other_balance'] - db_data['other_balance']) > 0.01:
            wallet_match = False
            wallet_issues.append(f"Other: Excel=₹{excel_data['other_balance']:.2f} vs DB=₹{db_data['other_balance']:.2f}")
        
        if not wallet_match:
            wallet_mismatches.append({
                'user': db_data['display_id'] or excel_data['email'],
                'name': db_data['name'],
                'excel': excel_data,
                'db': db_data,
                'issues': wallet_issues
            })
        
        # Check packages
        excel_pkg_list = excel_data['packages']
        db_pkg_list = db_data['packages']
        
        # Match packages by amount
        matched_packages = []
        unmatched_excel = []
        unmatched_db = []
        pkg_mismatches = []
        
        for excel_pkg in excel_pkg_list:
            matched = False
            for db_pkg in db_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    # Check for mismatches
                    issues = []
                    if abs(excel_pkg['income'] - db_pkg['income']) > 0.01:
                        issues.append(f"Income: Excel=₹{excel_pkg['income']:.2f} vs DB=₹{db_pkg['income']:.2f}")
                    if excel_pkg['used_global_id'] != db_pkg['effective_global_ids']:
                        issues.append(f"Global IDs: Excel={excel_pkg['used_global_id']} vs DB={db_pkg['effective_global_ids']}")
                    
                    if issues:
                        pkg_mismatches.append({
                            'user': db_data['display_id'] or excel_data['email'],
                            'name': db_data['name'],
                            'package': excel_pkg['name'],
                            'amount': excel_pkg['amount'],
                            'excel': excel_pkg,
                            'db': db_pkg,
                            'issues': issues
                        })
                    else:
                        matched_packages.append({
                            'excel': excel_pkg,
                            'db': db_pkg
                        })
                    break
            if not matched:
                unmatched_excel.append(excel_pkg)
        
        for db_pkg in db_pkg_list:
            matched = False
            for excel_pkg in excel_pkg_list:
                if abs(excel_pkg['amount'] - db_pkg['amount']) < 0.01:
                    matched = True
                    break
            if not matched:
                unmatched_db.append(db_pkg)
        
        if unmatched_excel or unmatched_db or pkg_mismatches:
            package_mismatches.append({
                'user': db_data['display_id'] or excel_data['email'],
                'name': db_data['name'],
                'matched': matched_packages,
                'unmatched_excel': unmatched_excel,
                'unmatched_db': unmatched_db,
                'pkg_mismatches': pkg_mismatches
            })
        
        if wallet_match and not unmatched_excel and not unmatched_db and not pkg_mismatches:
            matched_users.append({
                'user': db_data['display_id'] or excel_data['email'],
                'name': db_data['name'],
                'packages': len(matched_packages)
            })
    
    # Check DB users not in Excel
    for key, db_data in db_users.items():
        if key not in excel_users:
            missing_in_excel.append(db_data)
    
    # Print summary
    print(f"✅ Perfect Matches: {len(matched_users)}")
    print(f"⚠️  Wallet Mismatches: {len(wallet_mismatches)}")
    print(f"⚠️  Package Mismatches: {len(package_mismatches)}")
    print(f"❌ Missing in DB: {len(missing_in_db)}")
    print(f"❌ Missing in Excel: {len(missing_in_excel)}")
    print()
    
    # Print wallet mismatches (first 20)
    if wallet_mismatches:
        print("=" * 120)
        print("⚠️  WALLET MISMATCHES (showing first 20)")
        print("=" * 120)
        for mismatch in wallet_mismatches[:20]:
            print(f"\n👤 {mismatch['user']} - {mismatch['name']}")
            for issue in mismatch['issues']:
                print(f"   ⚠️  {issue}")
        if len(wallet_mismatches) > 20:
            print(f"\n... and {len(wallet_mismatches) - 20} more wallet mismatches")
    
    # Print package mismatches (first 20)
    if package_mismatches:
        print("\n" + "=" * 120)
        print("⚠️  PACKAGE MISMATCHES (showing first 20)")
        print("=" * 120)
        for mismatch in package_mismatches[:20]:
            print(f"\n👤 {mismatch['user']} - {mismatch['name']}")
            
            # Package data mismatches
            for pkg_mismatch in mismatch['pkg_mismatches']:
                print(f"   📦 {pkg_mismatch['package']} (₹{pkg_mismatch['amount']:,.2f})")
                for issue in pkg_mismatch['issues']:
                    print(f"      ⚠️  {issue}")
            
            # Unmatched Excel packages
            if mismatch['unmatched_excel']:
                print(f"   📦 Packages in Excel but NOT in DB:")
                for pkg in mismatch['unmatched_excel']:
                    print(f"      - {pkg['name']} (₹{pkg['amount']:,.2f})")
            
            # Unmatched DB packages
            if mismatch['unmatched_db']:
                print(f"   📦 Packages in DB but NOT in Excel:")
                for pkg in mismatch['unmatched_db']:
                    print(f"      - {pkg['name']} (₹{pkg['amount']:,.2f})")
        
        if len(package_mismatches) > 20:
            print(f"\n... and {len(package_mismatches) - 20} more package mismatches")
    
    # Print missing in DB (first 20)
    if missing_in_db:
        print("\n" + "=" * 120)
        print("❌ USERS IN EXCEL BUT NOT IN DB (showing first 20)")
        print("=" * 120)
        for user in missing_in_db[:20]:
            print(f"👤 {user['name']} ({user['email'] or user['member_id']})")
        if len(missing_in_db) > 20:
            print(f"\n... and {len(missing_in_db) - 20} more users")
    
    # Print missing in Excel (first 20)
    if missing_in_excel:
        print("\n" + "=" * 120)
        print("❌ USERS IN DB BUT NOT IN EXCEL (showing first 20)")
        print("=" * 120)
        for user in missing_in_excel[:20]:
            print(f"👤 {user['name']} ({user['display_id'] or user['email']})")
        if len(missing_in_excel) > 20:
            print(f"\n... and {len(missing_in_excel) - 20} more users")
    
    # Final summary
    print("\n" + "=" * 120)
    print("📊 FINAL SUMMARY")
    print("=" * 120)
    print(f"Total Excel Users: {excel_count}")
    print(f"Total DB Users: {db_count}")
    print(f"✅ Perfect Matches: {len(matched_users)}")
    print(f"⚠️  Wallet Issues: {len(wallet_mismatches)}")
    print(f"⚠️  Package Issues: {len(package_mismatches)}")
    print(f"❌ Missing in DB: {len(missing_in_db)}")
    print(f"❌ Missing in Excel: {len(missing_in_excel)}")
    
    match_percentage = (len(matched_users) / excel_count * 100) if excel_count > 0 else 0
    print(f"\n✅ Match Rate: {match_percentage:.2f}%")
    print("=" * 120)
    
    # Save detailed report
    report = {
        'verification_date': datetime.now().isoformat(),
        'summary': {
            'excel_users': excel_count,
            'db_users': db_count,
            'perfect_matches': len(matched_users),
            'wallet_mismatches': len(wallet_mismatches),
            'package_mismatches': len(package_mismatches),
            'missing_in_db': len(missing_in_db),
            'missing_in_excel': len(missing_in_excel),
            'match_percentage': match_percentage
        },
        'wallet_mismatches': wallet_mismatches[:100],
        'package_mismatches': package_mismatches[:100],
        'missing_in_db': missing_in_db[:100],
        'missing_in_excel': missing_in_excel[:100]
    }
    
    with open('all-users-verification-report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    print(f"\n📄 Detailed report saved to: all-users-verification-report.json")

if __name__ == '__main__':
    verify_all()

