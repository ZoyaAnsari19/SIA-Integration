#!/usr/bin/env python3
"""
Update all users' purchase income (self+global) from Excel
Matches by user and package amount, updates all matching purchases
"""

import openpyxl
import subprocess

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query, return_list=False):
    """Execute SQL query"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode == 0:
        if return_list:
            return [line.split('|') for line in result.stdout.strip().split('\n') if line.strip()]
        return result.stdout.strip()
    return None if not return_list else []

def update_purchase_income(purchase_id, income):
    """Update purchase income"""
    query = f"UPDATE purchases SET income = {income} WHERE id = {purchase_id};"
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    return result.returncode == 0

print("=" * 100)
print("💰 UPDATE ALL USERS INCOME (SELF+GLOBAL) FROM EXCEL")
print("=" * 100)
print()

# Read Excel
print("📄 Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

excel_user_packages = {}

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    
    if not email and not member_id:
        continue
    
    key_email = (email or '').lower()
    key_display = member_id
    
    packages = []
    for pack_num in range(1, 9):
        pack_amt_col = f'Pack. {pack_num} amt.'
        pack_income_col = f'Pack. {pack_num} self + global'
        
        pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
        pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else None
        
        if pack_amt and isinstance(pack_amt, (int, float)) and float(pack_amt) > 0:
            income_val = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0
            packages.append({
                'amount': float(pack_amt),
                'income': income_val
            })
    
    if packages:
        if key_email:
            excel_user_packages[key_email] = packages
        if key_display:
            excel_user_packages[key_display] = packages

print(f"✅ Found {len(excel_user_packages)} users with packages in Excel")
print()

# Get all DB purchases with user info
print("🗄️  Reading database purchases...")
query = """
SELECT 
    pur.id,
    LOWER(u.email) as email,
    u.display_id,
    pur.amount,
    pur.income
FROM purchases pur
JOIN users u ON pur.user_id = u.id
WHERE pur.status = 'completed'
ORDER BY u.id, pur.amount, pur.id;
"""

db_purchases = run_sql_query(query, return_list=True)
print(f"✅ Found {len(db_purchases)} purchases in DB")
print()

# Match and update
print("🔄 Matching and updating income...")
print()

updated = 0
not_matched = 0

for purchase in db_purchases:
    if len(purchase) < 5:
        continue
    
    purchase_id = purchase[0].strip()
    email = purchase[1].strip() if purchase[1] else None
    display_id = purchase[2].strip() if len(purchase) > 2 and purchase[2] else None
    amount = float(purchase[3].strip()) if purchase[3] else 0
    current_income = float(purchase[4].strip()) if len(purchase) > 4 and purchase[4] else 0
    
    # Find matching Excel package
    excel_packages = None
    if email:
        excel_packages = excel_user_packages.get(email)
    if not excel_packages and display_id:
        excel_packages = excel_user_packages.get(display_id)
    
    if not excel_packages:
        not_matched += 1
        continue
    
    # Find matching package by amount
    matching_pkg = None
    for pkg in excel_packages:
        if abs(pkg['amount'] - amount) < 0.01:  # Allow small floating point differences
            matching_pkg = pkg
            break
    
    if not matching_pkg:
        not_matched += 1
        continue
    
    excel_income = matching_pkg['income']
    
    # Update if different
    if abs(current_income - excel_income) > 0.01:
        if update_purchase_income(purchase_id, excel_income):
            updated += 1
            if updated <= 30:
                identifier = email or display_id or purchase_id
                print(f"  ✅ Purchase {purchase_id} ({identifier} - ₹{amount:,.2f}): {current_income:.2f} → {excel_income:.2f}")

print()
print("=" * 100)
print(f"✅ COMPLETE:")
print(f"   - Purchases updated: {updated}")
print(f"   - Purchases not matched: {not_matched}")
print(f"   - Total purchases processed: {len(db_purchases)}")
print("=" * 100)

