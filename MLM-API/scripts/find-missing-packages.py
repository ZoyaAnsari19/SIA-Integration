#!/usr/bin/env python3
"""
Find users in DB who have missing packages according to Excel
Match by package amount (not name)
"""

import openpyxl
import subprocess
import json

# Configuration
EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query via docker exec"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        print(f"Query error: {result.stderr[:200]}")
        return None
    
    return result.stdout.strip()

def get_excel_users_with_packages():
    """Read users with packages from Excel"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return {}
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    excel_users = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
        
        if not email and not member_id:
            continue
        
        # Get packages
        packages = []
        for pack_num in range(1, 9):
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_name_col = f'Pack. {pack_num} name'
            
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            if pack_amt is not None:
                try:
                    amount = float(pack_amt) if isinstance(pack_amt, (int, float)) else None
                    if amount and amount > 0:
                        pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
                        packages.append({
                            'amount': amount,
                            'name': str(pack_name) if pack_name else 'N/A'
                        })
                except:
                    pass
        
        if not packages:
            continue
        
        # Create key for matching (email or member_id)
        key = email.lower() if email else (member_id.upper() if member_id else None)
        if not key:
            continue
        
        excel_users[key] = {
            'email': email,
            'member_id': member_id,
            'name': name,
            'packages': packages
        }
    
    print(f"✅ Found {len(excel_users)} users with packages in Excel")
    return excel_users

def get_db_users_with_packages():
    """Get ALL users from database (with or without packages)"""
    print("🗄️  Reading database...")
    
    # First get all users
    all_users_query = """
    SELECT 
        u.id,
        u.display_id,
        u.email,
        u.name
    FROM users u
    ORDER BY u.id;
    """
    
    all_users_result = run_sql_query(all_users_query)
    if not all_users_result:
        print("❌ Failed to get users from database")
        return {}
    
    db_users = {}
    
    # Initialize all users
    for line in all_users_result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 4:
            user_id = parts[0].strip()
            display_id = parts[1].strip() if parts[1] else None
            email = parts[2].strip() if parts[2] else None
            name = parts[3].strip() if parts[3] else None
            
            # Create key for matching (email or display_id)
            key = email.lower() if email else (display_id.upper() if display_id else None)
            if not key:
                continue
            
            db_users[key] = {
                'id': user_id,
                'display_id': display_id,
                'email': email,
                'name': name,
                'packages': []
            }
    
    # Now get packages for users who have them
    packages_query = """
    SELECT 
        u.id,
        u.email,
        u.display_id,
        pkg.id as package_id,
        pkg.name as package_name,
        pkg.price as package_amount,
        pur.id as purchase_id
    FROM users u
    JOIN purchases pur ON pur.user_id = u.id
    JOIN packages pkg ON pur.package_id = pkg.id
    ORDER BY u.id, pur.purchased_at;
    """
    
    packages_result = run_sql_query(packages_query)
    if packages_result:
        for line in packages_result.split('\n'):
            if not line.strip():
                continue
            parts = line.split('|')
            if len(parts) >= 7:
                user_id = parts[0].strip()
                email = parts[1].strip() if parts[1] else None
                display_id = parts[2].strip() if parts[2] else None
                package_id = parts[3].strip() if parts[3] else None
                package_name = parts[4].strip() if parts[4] else None
                package_amount = float(parts[5].strip()) if parts[5] else 0.0
                purchase_id = parts[6].strip() if parts[6] else None
                
                # Create key for matching
                key = email.lower() if email else (display_id.upper() if display_id else None)
                if not key or key not in db_users:
                    continue
                
                if package_id and package_amount > 0:
                    db_users[key]['packages'].append({
                        'amount': package_amount,
                        'name': package_name or 'N/A',
                        'package_id': package_id,
                        'purchase_id': purchase_id
                    })
    
    users_with_packages = sum(1 for u in db_users.values() if u['packages'])
    print(f"✅ Found {len(db_users)} total users in database ({users_with_packages} with packages)")
    return db_users

def find_missing_packages():
    """Find users with missing packages"""
    print("=" * 100)
    print("🔍 FINDING USERS WITH MISSING PACKAGES")
    print("=" * 100)
    print()
    
    excel_users = get_excel_users_with_packages()
    db_users = get_db_users_with_packages()
    
    print()
    print("=" * 100)
    print("📊 USERS WITH MISSING PACKAGES IN DB")
    print("=" * 100)
    print()
    
    missing_packages_list = []
    
    # Check each Excel user
    for key, excel_data in excel_users.items():
        db_data = db_users.get(key)
        
        # Skip if user doesn't exist in DB
        if not db_data:
            continue
        
        # Skip if user has no packages in Excel (shouldn't happen but check anyway)
        if not excel_data['packages']:
            continue
        
        excel_pkg_amounts = {pkg['amount'] for pkg in excel_data['packages']}
        db_pkg_amounts = {pkg['amount'] for pkg in db_data['packages']}
        
        # Find missing packages (in Excel but not in DB)
        missing_amounts = excel_pkg_amounts - db_pkg_amounts
        
        if missing_amounts:
            missing_packages = [pkg for pkg in excel_data['packages'] if pkg['amount'] in missing_amounts]
            missing_packages_list.append({
                'user_id': db_data['id'],
                'display_id': db_data['display_id'],
                'email': excel_data['email'] or db_data['email'],
                'name': excel_data['name'] or db_data['name'],
                'excel_packages': excel_data['packages'],
                'db_packages': db_data['packages'],
                'missing_packages': missing_packages
            })
    
    # Print results
    if not missing_packages_list:
        print("✅ No users found with missing packages!")
        print("All packages in Excel are present in database.")
        return
    
    print(f"Found {len(missing_packages_list)} users with missing packages:\n")
    
    for i, user in enumerate(missing_packages_list, 1):
        print(f"{i}. {user['name'] or 'N/A'}")
        print(f"   Email: {user['email'] or 'N/A'}")
        print(f"   Display ID: {user['display_id'] or 'N/A'}")
        print(f"   User ID: {user['user_id']}")
        print(f"   Missing Packages ({len(user['missing_packages'])}):")
        for pkg in user['missing_packages']:
            print(f"      - {pkg['name']}: ₹{pkg['amount']:.2f}")
        print(f"   Excel Packages ({len(user['excel_packages'])}):")
        for pkg in user['excel_packages']:
            status = "❌ MISSING" if pkg['amount'] in {mp['amount'] for mp in user['missing_packages']} else "✅"
            print(f"      {status} {pkg['name']}: ₹{pkg['amount']:.2f}")
        print(f"   DB Packages ({len(user['db_packages'])}):")
        for pkg in user['db_packages']:
            print(f"      ✅ {pkg['name']}: ₹{pkg['amount']:.2f}")
        print()
    
    print("=" * 100)
    print("📈 SUMMARY")
    print("=" * 100)
    print(f"Total users with missing packages: {len(missing_packages_list)}")
    total_missing = sum(len(u['missing_packages']) for u in missing_packages_list)
    print(f"Total missing packages: {total_missing}")
    print()
    
    return missing_packages_list

if __name__ == '__main__':
    find_missing_packages()

