#!/usr/bin/env python3
"""
Update purchase income from Excel data for recently added packages
"""

import openpyxl
import subprocess
import re

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode == 0:
        return result.stdout.strip()
    return None

def update_purchase_income(purchase_id, income):
    """Update purchase income"""
    query = f"UPDATE purchases SET income = {income} WHERE id = {purchase_id};"
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    return result.returncode == 0

print('📊 Reading Excel income data and updating purchases...')
print('=' * 80)

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# Find header row
header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

if not header_row:
    print("❌ Header row not found")
    exit(1)

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

purchase_updates = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    
    if not email and not member_id:
        continue
    
    # Get packages from Excel
    for pack_num in range(1, 9):
        pack_amt_col = f'Pack. {pack_num} amt.'
        pack_income_col = f'Pack. {pack_num} self + global'
        
        pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
        pack_income = ws.cell(row_idx, headers.get(pack_income_col, 0)).value if headers.get(pack_income_col) else None
        
        if pack_amt and isinstance(pack_amt, (int, float)) and float(pack_amt) > 0:
            income_val = float(pack_income) if pack_income and isinstance(pack_income, (int, float)) else 0
            
            # Find matching purchase in DB (only for recently added purchases)
            if email:
                query = f"SELECT pur.id FROM purchases pur JOIN users u ON pur.user_id = u.id WHERE LOWER(u.email) = LOWER('{email}') AND pur.amount = {float(pack_amt)} AND pur.id >= 1522 ORDER BY pur.id DESC LIMIT 1;"
            else:
                query = f"SELECT pur.id FROM purchases pur JOIN users u ON pur.user_id = u.id WHERE u.display_id = '{member_id}' AND pur.amount = {float(pack_amt)} AND pur.id >= 1522 ORDER BY pur.id DESC LIMIT 1;"
            
            purchase_id = run_sql_query(query)
            
            if purchase_id and purchase_id.isdigit():
                purchase_updates.append((purchase_id, income_val, email or member_id, pack_amt))

print(f'Found {len(purchase_updates)} purchases to update')
print()

# Update income
updated = 0
for purchase_id, income, identifier, amount in purchase_updates:
    if update_purchase_income(purchase_id, income):
        print(f'✅ Purchase {purchase_id} ({identifier} - ₹{amount}): Income = ₹{income:.2f}')
        updated += 1
    else:
        print(f'❌ Failed to update purchase {purchase_id}')

print()
print(f'✅ Updated {updated} purchases with income data from Excel')

