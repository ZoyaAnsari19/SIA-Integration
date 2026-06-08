#!/usr/bin/env python3
"""
Update user wallet balances (spot and other) from Excel data
"""

import openpyxl
import subprocess

EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query):
    """Execute SQL query"""
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -c '{escaped_query}'"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    return result.returncode == 0

def update_wallet(user_id, spot_balance, other_balance):
    """Update user wallet balances"""
    total_balance = spot_balance + other_balance
    
    query = f"""
    INSERT INTO user_balances (
      user_id,
      spot_balance,
      other_balance,
      balance
    ) VALUES (
      {user_id},
      {spot_balance},
      {other_balance},
      {total_balance}
    )
    ON CONFLICT (user_id) DO UPDATE SET
      spot_balance = {spot_balance},
      other_balance = {other_balance},
      balance = {total_balance};
    """
    
    return run_sql_query(query)

def get_user_id(display_id, email):
    """Get user ID from display_id or email"""
    if display_id:
        query = f"SELECT id FROM users WHERE display_id = '{display_id}' LIMIT 1;"
        escaped_query = query.replace("'", "'\\''")
        cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -c '{escaped_query}'"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    
    if email:
        query = f"SELECT id FROM users WHERE LOWER(email) = LOWER('{email}') LIMIT 1;"
        escaped_query = query.replace("'", "'\\''")
        cmd = ['docker', 'exec', LOCAL_CONTAINER, 'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -c '{escaped_query}'"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    
    return None

print("=" * 100)
print("💰 UPDATE WALLET BALANCES FROM EXCEL")
print("=" * 100)
print()

# Read Excel
print("📄 Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

header_row = None
for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
    if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
        header_row = row_idx
        break

headers = {}
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

users_to_update = []

for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
    email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
    member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
    name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
    
    if not email and not member_id:
        continue
    
    spot_amt = ws.cell(row_idx, headers.get('Spot amt.', 0)).value if headers.get('Spot amt.') else 0
    other_amt = ws.cell(row_idx, headers.get('Other amt.', 0)).value if headers.get('Other amt.') else 0
    
    spot = float(spot_amt) if spot_amt and isinstance(spot_amt, (int, float)) else 0
    other = float(other_amt) if other_amt and isinstance(other_amt, (int, float)) else 0
    
    users_to_update.append({
        'email': email,
        'display_id': member_id,
        'name': name,
        'spot': spot,
        'other': other
    })

print(f"✅ Found {len(users_to_update)} users in Excel")
print()

# Update wallets
updated = 0
not_found = 0
errors = 0

print("🔄 Updating wallets...")
print()

for user_data in users_to_update:
    user_id = get_user_id(user_data['display_id'], user_data['email'])
    
    if not user_id:
        not_found += 1
        if not_found <= 10:  # Show first 10 not found
            print(f"  ⚠️  User not found: {user_data['name'] or user_data['display_id'] or user_data['email']}")
        continue
    
    if update_wallet(user_id, user_data['spot'], user_data['other']):
        updated += 1
        if updated <= 20:  # Show first 20 updates
            print(f"  ✅ {user_data['name'] or user_data['display_id'] or user_data['email']}: Spot ₹{user_data['spot']:,.2f}, Other ₹{user_data['other']:,.2f}")
    else:
        errors += 1
        if errors <= 10:  # Show first 10 errors
            print(f"  ❌ Failed to update: {user_data['name'] or user_data['display_id'] or user_data['email']}")

print()
print("=" * 100)
print(f"✅ COMPLETE:")
print(f"   - Updated: {updated} users")
print(f"   - Not found: {not_found} users")
print(f"   - Errors: {errors} users")
print("=" * 100)

