#!/usr/bin/env python3
"""
Add missing packages to DB for users who have packages in Excel but not in DB
Set global_ids to match Excel
"""

import openpyxl
import subprocess
import json
from datetime import datetime, timedelta

# Configuration
EXCEL_FILE = '/Users/faizanansari/Documents/MLM-bilal-sir/MLM/products-export-3.xlsx'
LOCAL_CONTAINER = 'postgres0-local-db'
DB_USER = 'mlm_user'
DB_NAME = 'mlm_commission'

def run_sql_query(query, description="", return_id=False):
    """Execute SQL query via docker exec"""
    if description:
        print(f"    {description}...")
    
    query = ' '.join(query.split())
    escaped_query = query.replace("'", "'\\''")
    
    cmd = [
        'docker', 'exec', LOCAL_CONTAINER,
        'bash', '-c', f"psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_query}'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        if description:
            error_msg = result.stderr[:200] if result.stderr else "Unknown error"
            print(f"    ❌ Error: {error_msg}")
        return None
    
    output = result.stdout.strip()
    
    if return_id and output:
        lines = output.split('\n')
        for line in lines:
            line = line.strip()
            if line and not line.startswith('INSERT') and line.isdigit():
                return line
        numbers = re.findall(r'\d+', output)
        if numbers:
            return numbers[-1]
    
    return output

def get_users_with_missing_packages():
    """Get users who have packages in Excel but not in DB"""
    print("📄 Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
        if any(cell.value and 'ID' in str(cell.value).upper() for cell in row):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Header row not found")
        return []
    
    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    users_with_missing = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
        email = ws.cell(row_idx, headers.get('Email', 0)).value if headers.get('Email') else None
        member_id = ws.cell(row_idx, headers.get('Member ID', 0)).value if headers.get('Member ID') else None
        name = ws.cell(row_idx, headers.get('Name', 0)).value if headers.get('Name') else None
        
        if not email and not member_id:
            continue
        
        # Get packages from Excel
        packages = []
        for pack_num in range(1, 9):
            pack_amt_col = f'Pack. {pack_num} amt.'
            pack_name_col = f'Pack. {pack_num} name'
            pack_global_col = f'Pack. {pack_num} used global id'
            
            pack_amt = ws.cell(row_idx, headers.get(pack_amt_col, 0)).value if headers.get(pack_amt_col) else None
            if pack_amt is not None:
                try:
                    amount = float(pack_amt) if isinstance(pack_amt, (int, float)) else None
                    if amount and amount > 0:
                        pack_name = ws.cell(row_idx, headers.get(pack_name_col, 0)).value if headers.get(pack_name_col) else 'N/A'
                        pack_global = ws.cell(row_idx, headers.get(pack_global_col, 0)).value if headers.get(pack_global_col) else 0
                        used_global_id = int(pack_global) if pack_global and isinstance(pack_global, (int, float)) else 0
                        
                        packages.append({
                            'amount': amount,
                            'name': str(pack_name) if pack_name else 'N/A',
                            'used_global_id': used_global_id
                        })
                except:
                    pass
        
        if not packages:
            continue
        
        # Check if user exists in DB
        key = email.lower() if email else (member_id.upper() if member_id else None)
        if not key:
            continue
        
        # Get user from DB
        if email:
            escaped_email = email.replace("'", "''")
            query = f"SELECT id, email, display_id, name FROM users WHERE LOWER(email) = LOWER('{escaped_email}') LIMIT 1;"
        else:
            escaped_member_id = member_id.replace("'", "''")
            query = f"SELECT id, email, display_id, name FROM users WHERE LOWER(display_id) = LOWER('{escaped_member_id}') LIMIT 1;"
        
        user_result = run_sql_query(query)
        if not user_result:
            continue
        
        parts = user_result.split('|')
        if len(parts) < 4:
            continue
        
        user_id = parts[0].strip()
        db_email = parts[1].strip() if parts[1] else None
        db_display_id = parts[2].strip() if parts[2] else None
        db_name = parts[3].strip() if parts[3] else None
        
        # Check existing purchases for this user
        purchase_query = f"SELECT package_id, amount FROM purchases WHERE user_id = {user_id};"
        purchase_result = run_sql_query(purchase_query)
        
        existing_package_amounts = set()
        if purchase_result:
            for line in purchase_result.split('\n'):
                if not line.strip():
                    continue
                parts = line.split('|')
                if len(parts) >= 2:
                    try:
                        pkg_amount = float(parts[1].strip()) if parts[1] else 0.0
                        if pkg_amount > 0:
                            existing_package_amounts.add(pkg_amount)
                    except:
                        pass
        
        # Find missing packages
        missing_packages = []
        for pkg in packages:
            # Check if package exists in DB (by amount)
            found = False
            for existing_amt in existing_package_amounts:
                if abs(pkg['amount'] - existing_amt) < 0.01:
                    found = True
                    break
            if not found:
                missing_packages.append(pkg)
        
        if missing_packages:
            users_with_missing.append({
                'user_id': user_id,
                'email': db_email or email,
                'display_id': db_display_id or member_id,
                'name': db_name or name,
                'missing_packages': missing_packages
            })
    
    print(f"✅ Found {len(users_with_missing)} users with missing packages")
    return users_with_missing

def find_package_by_amount(amount):
    """Find package_id by amount"""
    query = f"SELECT id FROM packages WHERE ABS(price - {amount}) < 0.01 LIMIT 1;"
    result = run_sql_query(query)
    if result:
        return result.strip()
    return None

def check_if_reinvestment(user_id, purchase_date):
    """Check if this is a reinvestment"""
    # Check if user has purchases before this date
    query = f"""
    SELECT COUNT(*) 
    FROM purchases 
    WHERE user_id = {user_id} 
    AND purchased_at < '{purchase_date}'::timestamp
    AND status = 'completed';
    """
    result = run_sql_query(query)
    if result:
        try:
            count = int(result.strip())
            return count > 0
        except:
            pass
    return False

def check_eligibility(user_id, level):
    """Check if user is eligible for a specific level"""
    query = f"SELECT eligibility::text FROM level_eligibility WHERE user_id = {user_id};"
    result = run_sql_query(query)
    if result:
        try:
            eligibility = json.loads(result)
            return bool(eligibility.get(str(level), False))
        except:
            try:
                cleaned = result.strip().strip('"').replace("\\\"", '"')
                eligibility = json.loads(cleaned)
                return bool(eligibility.get(str(level), False))
            except:
                pass
    return False

def get_spot_percent(level):
    """Get SPOT commission percentage for a level"""
    query = f"SELECT spot_commission_percent FROM levels WHERE level = {level};"
    result = run_sql_query(query)
    if result:
        try:
            return float(result.strip())
        except:
            pass
    return 0.0

def get_uplines(user_id, max_depth=10):
    """Get all uplines for a user"""
    query = f"""
    SELECT 
        ancestor_id,
        depth
    FROM user_tree_paths
    WHERE descendant_id = {user_id}
    AND depth BETWEEN 1 AND {max_depth}
    ORDER BY depth;
    """
    
    result = run_sql_query(query)
    if not result:
        return []
    
    uplines = []
    for line in result.split('\n'):
        if not line.strip():
            continue
        parts = line.split('|')
        if len(parts) >= 2:
            uplines.append({
                'ancestor_id': parts[0].strip(),
                'depth': int(parts[1].strip()) if parts[1] else 0
            })
    
    return uplines

def process_spot_commissions(buyer_id, purchase_id, purchase_amount, is_reinvestment):
    """Process SPOT commissions for uplines (unqualified only)"""
    uplines = get_uplines(buyer_id, max_depth=10)
    
    pending_count = 0
    skipped_qualified = 0
    skipped_direct = 0
    
    for upline in uplines:
        depth = upline['depth']
        level = depth - 1  # depth 2 → level 1, etc.
        upline_id = upline['ancestor_id']
        
        # Skip direct referrer (Level 0, depth 1)
        if depth == 1:
            skipped_direct += 1
            continue
        
        # For Team Uplines (Level 1-9, depth 2-10)
        if depth >= 2 and depth <= 10:
            # Get SPOT percentage
            spot_percent = get_spot_percent(level)
            if spot_percent == 0:
                continue
            
            # Calculate SPOT amount based on invested amount
            spot_amount = purchase_amount * spot_percent / 100
            
            # Apply 50% reduction for reinvestments (Level 1+)
            if is_reinvestment and level >= 1:
                spot_amount = spot_amount * 0.5
            
            # Check eligibility
            eligible = check_eligibility(upline_id, level)
            
            if eligible:
                # Skip qualified upline - NO SPOT, NO pending
                skipped_qualified += 1
                continue
            else:
                # Unqualified upline - Add SPOT to pending_commissions
                pending_query = f"""
                INSERT INTO pending_commissions (
                  receiver_user_id,
                  source_user_id,
                  purchase_id,
                  level,
                  commission_type,
                  amount,
                  metadata
                ) VALUES (
                  {upline_id},
                  {buyer_id},
                  {purchase_id},
                  {level},
                  'SPOT',
                  {spot_amount},
                  '{{"level": {level}, "depth": {depth}, "reason": "eligibility", "is_reinvestment": {str(is_reinvestment).lower()}}}'::jsonb
                )
                ON CONFLICT DO NOTHING;
                """
                
                result = run_sql_query(pending_query)
                if result is not None:
                    pending_count += 1
    
    return {
        'pending_count': pending_count,
        'skipped_qualified': skipped_qualified,
        'skipped_direct': skipped_direct
    }

def create_purchase(user_id, package_id, amount, used_global_id, package_global_ids):
    """Create purchase record with global_ids"""
    # Calculate expiry (2x amount)
    expiry_days = int((amount * 2) / (amount * 0.02 / 30)) if amount > 0 else 365  # Approximate
    purchased_at = datetime.now()
    active_until = purchased_at + timedelta(days=expiry_days)
    
    # Set effective_global_ids to used_global_id from Excel
    effective_global_ids = used_global_id if used_global_id > 0 else None
    
    query = f"""
    INSERT INTO purchases (
      user_id,
      package_id,
      amount,
      purchased_at,
      active_until,
      status,
      effective_global_ids
    ) VALUES (
      {user_id},
      {package_id},
      {amount},
      '{purchased_at.isoformat()}',
      '{active_until.isoformat()}',
      'completed',
      {effective_global_ids if effective_global_ids else 'NULL'}
    )
    RETURNING id;
    """
    
    result = run_sql_query(query, f"Creating purchase for package ₹{amount}", return_id=True)
    return result

def main():
    import re
    
    print("=" * 100)
    print("🚀 ADDING MISSING PACKAGES TO DATABASE")
    print("=" * 100)
    print()
    
    # Get users with missing packages
    users_with_missing = get_users_with_missing_packages()
    
    if not users_with_missing:
        print("✅ No users with missing packages found!")
        return
    
    print()
    print("=" * 100)
    print("📦 PROCESSING MISSING PACKAGES")
    print("=" * 100)
    print()
    
    total_purchases = 0
    total_pending = 0
    
    for user_data in users_with_missing:
        user_id = user_data['user_id']
        user_name = user_data['name'] or user_data['email'] or user_data['display_id']
        
        print(f"\n👤 User: {user_name} (ID: {user_id})")
        print(f"   Missing Packages: {len(user_data['missing_packages'])}")
        
        for pkg in user_data['missing_packages']:
            print(f"\n   📦 Package: {pkg['name']} - ₹{pkg['amount']:.2f}")
            print(f"      Excel Global IDs: {pkg['used_global_id']}")
            
            # Find package_id
            package_id = find_package_by_amount(pkg['amount'])
            if not package_id:
                print(f"      ❌ Package not found in DB for amount ₹{pkg['amount']:.2f}")
                continue
            
            # Get package global_ids
            pkg_query = f"SELECT global_ids FROM packages WHERE id = {package_id};"
            pkg_result = run_sql_query(pkg_query)
            package_global_ids = int(pkg_result.strip()) if pkg_result else 0
            
            print(f"      Package ID: {package_id}, Total Global IDs: {package_global_ids}")
            
            # Check if purchase already exists
            existing_query = f"""
            SELECT id FROM purchases 
            WHERE user_id = {user_id} 
            AND package_id = {package_id}
            AND ABS(amount - {pkg['amount']}) < 0.01
            LIMIT 1;
            """
            existing = run_sql_query(existing_query)
            if existing:
                print(f"      ⚠️  Purchase already exists, skipping")
                continue
            
            # Create purchase
            purchase_id = create_purchase(
                user_id=user_id,
                package_id=package_id,
                amount=pkg['amount'],
                used_global_id=pkg['used_global_id'],
                package_global_ids=package_global_ids
            )
            
            if not purchase_id:
                print(f"      ❌ Failed to create purchase")
                continue
            
            print(f"      ✅ Purchase created: ID {purchase_id}")
            print(f"      ✅ Effective Global IDs set to: {pkg['used_global_id']}")
            total_purchases += 1
            
            # Check if reinvestment
            purchase_date = datetime.now().isoformat()
            is_reinvestment = check_if_reinvestment(user_id, purchase_date)
            
            if is_reinvestment:
                print(f"      🔄 Reinvestment detected")
            
            # Process SPOT commissions
            print(f"      💵 Processing SPOT commissions...")
            spot_result = process_spot_commissions(
                buyer_id=user_id,
                purchase_id=purchase_id,
                purchase_amount=pkg['amount'],
                is_reinvestment=is_reinvestment
            )
            
            print(f"         - Pending commissions: {spot_result['pending_count']}")
            print(f"         - Skipped (qualified): {spot_result['skipped_qualified']}")
            print(f"         - Skipped (direct): {spot_result['skipped_direct']}")
            total_pending += spot_result['pending_count']
    
    print()
    print("=" * 100)
    print("📈 SUMMARY")
    print("=" * 100)
    print(f"✅ Purchases Created: {total_purchases}")
    print(f"✅ Pending Commissions Created: {total_pending}")
    print()

if __name__ == '__main__':
    main()

